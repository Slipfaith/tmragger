from __future__ import annotations

import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from core.output_paths import sibling_output_dir


_XML_LANG_ATTR = "{http://www.w3.org/XML/1998/namespace}lang"


@dataclass(slots=True)
class ExcelToTmxResult:
    input_file: Path
    output_file: Path
    source_lang: str
    target_lang: str
    rows_read: int
    rows_written: int


def convert_excel_to_tmx(
    input_path: Path,
    source_lang: str,
    target_lang: str,
    has_header: bool = True,
    source_column: int = 1,
    target_column: int = 2,
    comment_column: int = 3,
    output_dir: Path | None = None,
) -> ExcelToTmxResult:
    source_lang = source_lang.strip()
    target_lang = target_lang.strip()
    if not source_lang or not target_lang:
        raise ValueError("Source and target language codes are required.")

    source_index = _column_to_index(source_column, "source")
    target_index = _column_to_index(target_column, "target")
    comment_index = _column_to_index(comment_column, "comment")
    if len({source_index, target_index, comment_index}) < 3:
        raise ValueError("Source, target and comment columns must be different.")

    resolved_output_dir = output_dir or sibling_output_dir(input_path)
    resolved_output_dir.mkdir(parents=True, exist_ok=True)
    output_file = resolved_output_dir / f"{input_path.stem}.tmx"

    tmx = ET.Element("tmx", {"version": "1.4"})
    ET.SubElement(
        tmx,
        "header",
        {
            "creationtool": "tmx2csv_app",
            "creationtoolversion": "1.0",
            "segtype": "sentence",
            "adminlang": "en-US",
            "srclang": source_lang,
            "datatype": "plaintext",
        },
    )
    body = ET.SubElement(tmx, "body")

    rows_read = 0
    rows_written = 0

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        row_iter = sheet.iter_rows(values_only=True)
        if has_header:
            next(row_iter, None)

        for row in row_iter:
            rows_read += 1
            values = tuple(row or ())
            source_text = _cell_to_text(values, source_index)
            target_text = _cell_to_text(values, target_index)
            comment_text = _cell_to_text(values, comment_index)
            if not (source_text or target_text or comment_text):
                continue

            tu = ET.SubElement(body, "tu")
            _add_memoq_tu_props(tu, comment_text)
            source_tuv = ET.SubElement(tu, "tuv", {_XML_LANG_ATTR: source_lang})
            source_seg = ET.SubElement(source_tuv, "seg")
            source_seg.text = source_text

            target_tuv = ET.SubElement(tu, "tuv", {_XML_LANG_ATTR: target_lang})
            target_seg = ET.SubElement(target_tuv, "seg")
            target_seg.text = target_text
            rows_written += 1
    finally:
        workbook.close()

    ET.indent(tmx, space="  ")
    tree = ET.ElementTree(tmx)
    tree.write(output_file, encoding="utf-8", xml_declaration=True)

    return ExcelToTmxResult(
        input_file=input_path,
        output_file=output_file,
        source_lang=source_lang,
        target_lang=target_lang,
        rows_read=rows_read,
        rows_written=rows_written,
    )


def _cell_to_text(row: tuple[object, ...], index: int) -> str:
    if index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value)


def _column_to_index(column: int, name: str) -> int:
    if column < 1:
        raise ValueError(f"{name.capitalize()} column must be >= 1.")
    return column - 1


def _add_memoq_tu_props(tu: ET.Element, comment_text: str) -> None:
    ET.SubElement(tu, "prop", {"type": "client"}).text = " "
    ET.SubElement(tu, "prop", {"type": "project"}).text = " "
    ET.SubElement(tu, "prop", {"type": "domain"}).text = " "
    ET.SubElement(tu, "prop", {"type": "subject"}).text = " "
    ET.SubElement(tu, "prop", {"type": "corrected"}).text = "no"
    ET.SubElement(tu, "prop", {"type": "aligned"}).text = "no"
    ET.SubElement(tu, "prop", {"type": "x-Comment"}).text = comment_text or " "
