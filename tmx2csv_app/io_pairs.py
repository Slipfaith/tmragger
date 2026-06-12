from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from openpyxl import Workbook, load_workbook


@dataclass(slots=True)
class PairFileInfo:
    file_path: Path
    file_format: str
    source_lang: str
    target_lang: str
    row_count: int


@dataclass(slots=True)
class PairRow:
    row_index: int
    source: str
    target: str


def analyze_pair_file(path: Path) -> PairFileInfo:
    file_format = _detect_format(path)
    if file_format == "csv":
        with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            reader = csv.reader(file_obj)
            header = next(reader, None)
            source_lang, target_lang = _extract_header(path, header)
            row_count = sum(1 for _ in reader)
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            row_iter = sheet.iter_rows(values_only=True)
            header = next(row_iter, None)
            source_lang, target_lang = _extract_header(path, header)
            row_count = sum(1 for _ in row_iter)
        finally:
            workbook.close()

    return PairFileInfo(
        file_path=path,
        file_format=file_format,
        source_lang=source_lang,
        target_lang=target_lang,
        row_count=row_count,
    )


def iter_pair_rows(path: Path) -> Iterator[PairRow]:
    file_format = _detect_format(path)
    if file_format == "csv":
        with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            reader = csv.reader(file_obj)
            next(reader, None)
            for row_index, row in enumerate(reader, start=2):
                yield PairRow(row_index=row_index, source=_cell_to_text(row, 0), target=_cell_to_text(row, 1))
        return

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        row_iter = sheet.iter_rows(values_only=True)
        next(row_iter, None)
        for row_index, row in enumerate(row_iter, start=2):
            values = tuple(row or ())
            yield PairRow(row_index=row_index, source=_cell_value_to_text(values, 0), target=_cell_value_to_text(values, 1))
    finally:
        workbook.close()


def build_cleaned_output_path(input_path: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"{input_path.stem}__cleaned{input_path.suffix.lower()}"


class PairFileWriter:
    def write_row(self, source: str, target: str) -> None:
        raise NotImplementedError

    def close(self) -> None:
        raise NotImplementedError


class CsvPairWriter(PairFileWriter):
    def __init__(self, output_path: Path, source_lang: str, target_lang: str) -> None:
        self.output_path = output_path
        self._file = output_path.open("w", encoding="utf-8-sig", newline="")
        self._writer = csv.writer(self._file)
        self._writer.writerow([source_lang, target_lang])

    def write_row(self, source: str, target: str) -> None:
        self._writer.writerow([source, target])

    def close(self) -> None:
        self._file.close()


class XlsxPairWriter(PairFileWriter):
    def __init__(self, output_path: Path, source_lang: str, target_lang: str) -> None:
        self.output_path = output_path
        self._workbook = Workbook(write_only=True)
        self._sheet = self._workbook.create_sheet(title="Pairs")
        self._sheet.append([source_lang, target_lang])
        self._closed = False

    def write_row(self, source: str, target: str) -> None:
        self._sheet.append([source, target])

    def close(self) -> None:
        if self._closed:
            return
        self._workbook.save(self.output_path)
        self._closed = True


def open_pair_writer(output_path: Path, source_lang: str, target_lang: str) -> PairFileWriter:
    file_format = _detect_format(output_path)
    if file_format == "csv":
        return CsvPairWriter(output_path, source_lang, target_lang)
    return XlsxPairWriter(output_path, source_lang, target_lang)


def _detect_format(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return "csv"
    if suffix == ".xlsx":
        return "xlsx"
    raise ValueError(f"Unsupported pair file format: {path.name}")


def _extract_header(path: Path, header: object) -> tuple[str, str]:
    if not header:
        raise ValueError(f"{path.name} is empty.")
    values = list(header)
    if len(values) < 2:
        raise ValueError(f"{path.name} must contain at least 2 columns.")
    source_lang = str(values[0] or "").strip()
    target_lang = str(values[1] or "").strip()
    if not source_lang or not target_lang:
        raise ValueError(f"{path.name} must have language codes in the first two header cells.")
    return source_lang, target_lang


def _cell_to_text(row: list[str], index: int) -> str:
    if index >= len(row):
        return ""
    return str(row[index])


def _cell_value_to_text(row: tuple[object, ...], index: int) -> str:
    if index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value)
