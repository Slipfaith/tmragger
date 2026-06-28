"""Export progress-callback tests (no large sample fixture required)."""

from __future__ import annotations

import re
import zipfile

from core.offline_package import (
    MANIFEST_NAME,
    REPORT_XLSX_NAME,
    SOURCE_NAME,
    STATE_NAME,
    export_tmrepair_package,
)
from core.plan import Proposal, RepairPlan, make_cleanup_proposal_id


def _plan_with_proposals(count: int) -> RepairPlan:
    proposals = [
        Proposal(
            proposal_id=make_cleanup_proposal_id(i, "normalize_spaces", 0),
            kind="cleanup",
            tu_index=i,
            rule="normalize_spaces",
            message="ASCII spaces normalized.",
            before_tgt="a  b",
            after_tgt="a b",
            original_src="x",
            original_tgt="a  b",
        )
        for i in range(count)
    ]
    return RepairPlan(input_path="in.tmx", total_tus=count, proposals=proposals)


def test_export_reports_monotonic_row_progress(tmp_path):
    source = tmp_path / "in.tmx"
    source.write_bytes(b"<tmx version='1.4'></tmx>")
    count = 450
    package_path = tmp_path / "out.tmrepair"

    calls: list[tuple[int, int]] = []
    export_tmrepair_package(
        package_path=package_path,
        input_tmx_path=source,
        plan=_plan_with_proposals(count),
        settings={},
        progress_callback=lambda done, total: calls.append((done, total)),
    )

    assert calls, "progress callback was never invoked"
    assert all(total == count for _, total in calls)
    done_values = [done for done, _ in calls]
    assert done_values == sorted(done_values)  # never goes backwards
    assert calls[-1] == (count, count)  # ends at 100%
    assert len(calls) >= 3  # batched updates, not a single jump

    assert package_path.exists()
    with zipfile.ZipFile(package_path) as archive:
        names = set(archive.namelist())
    assert {MANIFEST_NAME, STATE_NAME, SOURCE_NAME, REPORT_XLSX_NAME} <= names


def test_export_without_callback_still_succeeds(tmp_path):
    source = tmp_path / "in.tmx"
    source.write_bytes(b"<tmx version='1.4'></tmx>")
    package_path = tmp_path / "out.tmrepair"

    export_tmrepair_package(
        package_path=package_path,
        input_tmx_path=source,
        plan=_plan_with_proposals(3),
        settings={},
    )

    assert package_path.exists()


def test_export_xlsx_marks_empty_segments_and_enables_header_filter(tmp_path):
    from openpyxl import load_workbook

    source = tmp_path / "in.tmx"
    source.write_bytes(b"<tmx version='1.4'></tmx>")
    proposals = [
        Proposal(
            proposal_id=make_cleanup_proposal_id(0, "dedup_tu", 0),
            kind="cleanup",
            tu_index=0,
            rule="dedup_tu",
            message="dup",
            original_src="Привет",
            original_tgt="Salom",
        ),
        Proposal(
            proposal_id=make_cleanup_proposal_id(1, "dedup_tu", 0),
            kind="cleanup",
            tu_index=1,
            rule="dedup_tu",
            message="dup",
            original_src="",
            original_tgt="",
        ),
    ]
    plan = RepairPlan(input_path="in.tmx", total_tus=2, proposals=proposals)
    package_path = tmp_path / "p.tmrepair"
    export_tmrepair_package(
        package_path=package_path, input_tmx_path=source, plan=plan, settings={}
    )

    with zipfile.ZipFile(package_path) as archive:
        archive.extract(REPORT_XLSX_NAME, tmp_path)
    sheet = load_workbook(tmp_path / REPORT_XLSX_NAME).active

    # Header filter is set and stays usable on the locked sheet.
    assert sheet.auto_filter.ref is not None
    assert sheet.auto_filter.ref.startswith("A4:")
    assert sheet.protection.autoFilter is False

    rows = {
        (str(sheet.cell(r, 3).value), str(sheet.cell(r, 4).value))
        for r in range(5, sheet.max_row + 1)
    }
    assert ("Привет", "Salom") in rows
    assert ("(пусто)", "(пусто)") in rows


def test_export_xlsx_diff_runs_are_opaque(tmp_path):
    # Whole-deletion rows (dedup/garbage) render with the "deleted" font; a
    # transparent (00 alpha) color blanks them out in viewers. The deletion
    # color must be opaque so the struck-through text stays visible.
    import io

    from openpyxl import load_workbook

    source = tmp_path / "in.tmx"
    source.write_bytes(b"<tmx/>")
    plan = RepairPlan(
        input_path="in.tmx",
        total_tus=1,
        proposals=[
            Proposal(
                proposal_id=make_cleanup_proposal_id(0, "dedup_tu", 0),
                kind="cleanup",
                tu_index=0,
                rule="dedup_tu",
                message="dup",
                original_src="Оцените работу",
                original_tgt="Baho bering",
            )
        ],
    )
    package_path = tmp_path / "p.tmrepair"
    export_tmrepair_package(
        package_path=package_path, input_tmx_path=source, plan=plan, settings={}
    )

    xlsx_bytes = zipfile.ZipFile(package_path).read(REPORT_XLSX_NAME)
    sheet = load_workbook(io.BytesIO(xlsx_bytes)).active
    font = sheet.cell(row=5, column=3).font  # the dedup TU source cell
    assert font.strike is True
    color = font.color.rgb
    assert color.endswith("B91C1C")  # red
    assert not color.startswith("00")  # opaque, not transparent


def test_export_strips_illegal_control_chars(tmp_path):
    # Segment text carrying XML-illegal control chars otherwise yields a
    # workbook Excel reports as corrupt; they must be stripped before export.
    import io
    import xml.etree.ElementTree as ET

    source = tmp_path / "in.tmx"
    source.write_bytes(b"<tmx/>")
    plan = RepairPlan(
        input_path="in.tmx",
        total_tus=1,
        proposals=[
            Proposal(
                proposal_id=make_cleanup_proposal_id(0, "dedup_tu", 0),
                kind="cleanup",
                tu_index=0,
                rule="dedup_tu",
                message="dup",
                original_src="Оцените\x0bработу\x07поддержки",
                original_tgt="Baho\x0cbering\x1f",
            )
        ],
    )
    package_path = tmp_path / "p.tmrepair"
    export_tmrepair_package(
        package_path=package_path, input_tmx_path=source, plan=plan, settings={}
    )

    xlsx_bytes = zipfile.ZipFile(package_path).read(REPORT_XLSX_NAME)
    sheet_xml = (
        zipfile.ZipFile(io.BytesIO(xlsx_bytes))
        .read("xl/worksheets/sheet1.xml")
    )
    assert not re.search(rb"[\x00-\x08\x0b\x0c\x0e-\x1f]", sheet_xml)
    ET.fromstring(sheet_xml)  # parses cleanly

