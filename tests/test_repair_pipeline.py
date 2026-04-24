import json
from pathlib import Path
import threading
import time

from openpyxl import load_workbook

from core.gemini_client import GeminiIssue, GeminiVerificationResult
from core.repair import (
    RepairStats,
    _created_tu_count_after_replacements,
    _dedup_segment_pair_key,
    _should_write_resume_checkpoint,
    _should_collect_report_details,
    repair_tmx_file,
)
from core.reports.html import write_html_diff_report


def _read(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def _write_sample_tmx(path: Path) -> None:
    path.write_text(
        """<?xml version="1.0" encoding="utf-8"?>
<tmx version="1.4">
  <header srclang="en-US" adminlang="en-US" creationtool="test" creationtoolversion="1.0" datatype="xml"/>
  <body>
    <tu creationid="u1">
      <prop type="x-Note">n1</prop>
      <tuv xml:lang="en-US"><seg>Hello world. Next sentence!</seg></tuv>
      <tuv xml:lang="ru-RU"><seg>Privet mir. Sleduiushchee predlozhenie!</seg></tuv>
    </tu>
    <tu creationid="u2">
      <tuv xml:lang="en-US"><seg>Single sentence only</seg></tuv>
      <tuv xml:lang="ru-RU"><seg>Tolko odno predlozhenie</seg></tuv>
    </tu>
  </body>
</tmx>
""",
        encoding="utf-8",
    )


def _write_multi_split_tmx(path: Path, count: int = 4) -> None:
    tus: list[str] = []
    for idx in range(1, count + 1):
        tus.append(
            f"""
    <tu creationid="u{idx}">
      <tuv xml:lang="en-US"><seg>Alpha {idx}. Beta {idx}!</seg></tuv>
      <tuv xml:lang="ru-RU"><seg>\u0410\u043b\u044c\u0444\u0430 {idx}. \u0411\u0435\u0442\u0430 {idx}!</seg></tuv>
    </tu>"""
        )
    body = "".join(tus)
    path.write_text(
        f"""<?xml version="1.0" encoding="utf-8"?>
<tmx version="1.4">
  <header srclang="en-US" adminlang="en-US" creationtool="test" creationtoolversion="1.0" datatype="xml"/>
  <body>{body}
  </body>
</tmx>
""",
        encoding="utf-8",
    )


def test_created_tu_count_uses_replacement_lengths_without_rewalking_body():
    assert _created_tu_count_after_replacements(
        total_tus=5,
        replacement_map={
            1: [],
            2: [object(), object(), object()],
            4: [object()],
        },
    ) == 6


def test_dedup_segment_pair_key_is_compact_for_large_segments():
    source = "A" * 50_000
    target = "B" * 50_000

    key = _dedup_segment_pair_key(source, target)

    assert key == _dedup_segment_pair_key(source, target)
    assert key != _dedup_segment_pair_key(source, target + "!")
    assert len(repr(key)) < 220


def test_resume_checkpoint_throttle_limits_repeated_serialization():
    assert _should_write_resume_checkpoint(
        processed_since_checkpoint=49,
        checkpoint_every_tus=50,
        now=100.0,
        last_checkpoint_at=0.0,
        min_interval_seconds=5.0,
    ) is False
    assert _should_write_resume_checkpoint(
        processed_since_checkpoint=50,
        checkpoint_every_tus=50,
        now=104.0,
        last_checkpoint_at=100.0,
        min_interval_seconds=5.0,
    ) is False
    assert _should_write_resume_checkpoint(
        processed_since_checkpoint=50,
        checkpoint_every_tus=50,
        now=105.0,
        last_checkpoint_at=100.0,
        min_interval_seconds=5.0,
    ) is True


def test_apply_only_checks_split_for_selected_tus(monkeypatch):
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_selected_split.tmx"
    output_path = runtime_dir / "output_selected_split.tmx"
    _write_multi_split_tmx(input_path, count=4)

    call_counter = {"count": 0}
    from core.splitter import propose_aligned_split as _real_propose_aligned_split

    def _counted_propose(*args, **kwargs):
        call_counter["count"] += 1
        return _real_propose_aligned_split(*args, **kwargs)

    monkeypatch.setattr("core.repair.propose_aligned_split", _counted_propose)

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        accepted_split_ids={"split:1"},
        accepted_cleanup_ids=set(),
        enable_split_short_sentence_pair_guard=False,
    )

    assert call_counter["count"] == 1
    assert stats.split_tus == 1

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)


def test_apply_only_checks_cleanup_for_selected_tus(monkeypatch):
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_selected_cleanup.tmx"
    output_path = runtime_dir / "output_selected_cleanup.tmx"
    _write_multi_split_tmx(input_path, count=5)

    call_counter = {"count": 0}
    from core.tm_cleanup import analyze_and_clean_segments as _real_cleanup

    def _counted_cleanup(*args, **kwargs):
        call_counter["count"] += 1
        return _real_cleanup(*args, **kwargs)

    monkeypatch.setattr("core.repair.analyze_and_clean_segments", _counted_cleanup)

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        accepted_split_ids=set(),
        accepted_cleanup_ids={"cleanup:2:normalize_spaces:0"},
        enable_split=False,
    )

    assert call_counter["count"] == 1
    assert stats.total_tus == 5

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)


def test_repair_tmx_file_splits_aligned_tu():
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input.tmx"
    output_path = runtime_dir / "output.tmx"
    html_report_path = runtime_dir / "report.html"
    xlsx_report_path = runtime_dir / "report.xlsx"
    _write_sample_tmx(input_path)

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        enable_split_short_sentence_pair_guard=False,
        html_report_path=html_report_path,
        xlsx_report_path=xlsx_report_path,
    )

    assert isinstance(stats, RepairStats)
    assert stats.total_tus == 2
    assert stats.split_tus == 1
    assert stats.created_tus == 3
    assert stats.high_confidence_splits == 1
    assert stats.medium_confidence_splits == 0
    assert stats.gemini_checked == 0
    assert stats.gemini_rejected == 0

    content = _read(output_path)
    assert content.count("<tu ") == 3
    assert "Hello world." in content
    assert "Next sentence!" in content
    assert "Single sentence only" in content
    assert '<prop type="x-TMXRepair-Confidence">HIGH</prop>' in content
    assert html_report_path.exists()
    assert xlsx_report_path.exists()
    html_content = _read(html_report_path)
    assert "TMX Repair Diff Report" in html_content
    assert "Hello world. Next sentence!" in html_content
    assert "Next sentence!" in html_content

    workbook = load_workbook(xlsx_report_path)
    assert workbook.sheetnames == [
        "Summary",
        "Split Changes",
        "Auto Cleanup",
        "Warnings",
        "Gemini Checks",
    ]
    assert workbook["Summary"]["A2"].value == "Input TMX"
    workbook.close()

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)
    html_report_path.unlink(missing_ok=True)
    xlsx_report_path.unlink(missing_ok=True)


def test_repair_with_gemini_fail_rejects_split():
    class AlwaysFailVerifier:
        def verify_split(self, request):
            return GeminiVerificationResult(
                verdict="FAIL",
                issues=[
                    GeminiIssue(
                        severity="high",
                        issue_type="alignment",
                        message="bad mapping",
                        src_index=0,
                        tgt_index=0,
                        suggestion="keep original tu",
                    )
                ],
                summary="rejected",
            )

    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_fail.tmx"
    output_path = runtime_dir / "output_fail.tmx"
    report_path = runtime_dir / "report_fail.json"
    _write_sample_tmx(input_path)

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        verify_with_gemini=True,
        gemini_verifier=AlwaysFailVerifier(),
        enable_split_short_sentence_pair_guard=False,
        report_path=report_path,
    )

    assert stats.split_tus == 0
    assert stats.created_tus == 2
    assert stats.gemini_checked == 1
    assert stats.gemini_rejected == 1
    assert stats.medium_confidence_splits == 0

    content = _read(output_path)
    assert content.count("<tu ") == 2
    assert '<prop type="x-TMXRepair-Confidence">' not in content

    report = json.loads(_read(report_path))
    assert report["gemini_checked"] == 1
    assert report["gemini_rejected"] == 1
    assert report["items"][0]["verdict"] == "FAIL"

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)
    report_path.unlink(missing_ok=True)


def test_repair_with_gemini_ok_marks_medium_confidence():
    class AlwaysOkVerifier:
        def verify_split(self, request):
            return GeminiVerificationResult(
                verdict="OK",
                issues=[],
                summary="looks good",
            )

    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_ok.tmx"
    output_path = runtime_dir / "output_ok.tmx"
    report_path = runtime_dir / "report_ok.json"
    _write_sample_tmx(input_path)

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        verify_with_gemini=True,
        gemini_verifier=AlwaysOkVerifier(),
        enable_split_short_sentence_pair_guard=False,
        report_path=report_path,
    )

    assert stats.split_tus == 1
    assert stats.created_tus == 3
    assert stats.gemini_checked == 1
    assert stats.gemini_rejected == 0
    assert stats.high_confidence_splits == 0
    assert stats.medium_confidence_splits == 1

    content = _read(output_path)
    assert content.count("<tu ") == 3
    assert '<prop type="x-TMXRepair-Confidence">MEDIUM</prop>' in content

    report = json.loads(_read(report_path))
    assert report["items"][0]["verdict"] == "OK"

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)
    report_path.unlink(missing_ok=True)


def test_repair_passes_custom_prompt_template_to_verifier():
    class CapturePromptVerifier:
        def __init__(self) -> None:
            self.captured_prompt_template = None

        def verify_split(self, request, prompt_template=None):
            self.captured_prompt_template = prompt_template
            return GeminiVerificationResult(
                verdict="OK",
                issues=[],
                summary="ok",
            )

    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_prompt.tmx"
    output_path = runtime_dir / "output_prompt.tmx"
    _write_sample_tmx(input_path)

    verifier = CapturePromptVerifier()
    custom_prompt = "CUSTOM PROMPT TEMPLATE {SRC_LANG} -> {TGT_LANG}"
    repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        verify_with_gemini=True,
        gemini_verifier=verifier,
        enable_split_short_sentence_pair_guard=False,
        gemini_prompt_template=custom_prompt,
    )

    assert verifier.captured_prompt_template == custom_prompt

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)


def test_repair_emits_progress_and_token_usage():
    class UsageVerifier:
        def verify_split(self, request):
            return GeminiVerificationResult(
                verdict="OK",
                issues=[],
                summary="ok",
                prompt_tokens=111,
                completion_tokens=22,
                total_tokens=133,
            )

    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_usage.tmx"
    output_path = runtime_dir / "output_usage.tmx"
    report_path = runtime_dir / "report_usage.json"
    _write_sample_tmx(input_path)

    progress_events: list[dict[str, object]] = []
    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        verify_with_gemini=True,
        gemini_verifier=UsageVerifier(),
        enable_split_short_sentence_pair_guard=False,
        report_path=report_path,
        progress_callback=lambda payload: progress_events.append(dict(payload)),
    )

    assert stats.gemini_input_tokens == 111
    assert stats.gemini_output_tokens == 22
    assert stats.gemini_total_tokens == 133
    assert stats.gemini_estimated_cost_usd > 0

    event_names = {str(event.get("event", "")) for event in progress_events}
    assert "file_start" in event_names
    assert "tu_start" in event_names
    assert "gemini_result" in event_names
    assert "file_complete" in event_names

    report = json.loads(_read(report_path))
    assert report["gemini_input_tokens"] == 111
    assert report["gemini_output_tokens"] == 22
    assert report["gemini_total_tokens"] == 133
    assert report["gemini_estimated_cost_usd"] > 0

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)
    report_path.unlink(missing_ok=True)


def test_repair_runs_gemini_verification_in_parallel_when_enabled():
    class SlowVerifier:
        def __init__(self) -> None:
            self.calls = 0
            self.in_flight = 0
            self.max_in_flight = 0
            self._lock = threading.Lock()

        def verify_split(self, request, prompt_template=None):  # noqa: ANN001
            with self._lock:
                self.calls += 1
                self.in_flight += 1
                self.max_in_flight = max(self.max_in_flight, self.in_flight)
            try:
                time.sleep(0.06)
                return GeminiVerificationResult(
                    verdict="OK",
                    issues=[],
                    summary="ok",
                )
            finally:
                with self._lock:
                    self.in_flight -= 1

    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_parallel.tmx"
    output_path = runtime_dir / "output_parallel.tmx"
    _write_multi_split_tmx(input_path, count=4)

    verifier = SlowVerifier()
    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        verify_with_gemini=True,
        gemini_verifier=verifier,
        gemini_max_parallel=3,
        enable_split_short_sentence_pair_guard=False,
    )

    assert stats.gemini_checked == 4
    assert verifier.calls == 4
    assert verifier.max_in_flight >= 2
    assert stats.split_tus == 4
    assert stats.created_tus == 8

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)


def test_repair_marks_unavailable_gemini_as_pending_without_applying_split():
    class UnavailableVerifier:
        def verify_split(self, request):  # noqa: ANN001
            return GeminiVerificationResult(
                verdict="WARN",
                issues=[],
                summary="Gemini request failed",
            )

    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_pending.tmx"
    output_path = runtime_dir / "output_pending.tmx"
    report_path = runtime_dir / "report_pending.json"
    _write_sample_tmx(input_path)

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        verify_with_gemini=True,
        gemini_verifier=UnavailableVerifier(),
        enable_split_short_sentence_pair_guard=False,
        report_path=report_path,
    )

    assert stats.split_tus == 0
    assert stats.skipped_tus >= 1
    content = _read(output_path)
    assert content.count("<tu ") == 2
    assert '<prop type="x-TMXRepair-Confidence">' not in content

    report = json.loads(_read(report_path))
    pending = report.get("pending_verification_events", [])
    assert isinstance(pending, list)
    assert len(pending) == 1
    assert "Gemini request failed" in str(pending[0].get("reason", ""))

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)
    report_path.unlink(missing_ok=True)


def test_repair_reuses_persistent_gemini_cache_between_runs():
    class CountingVerifier:
        def __init__(self) -> None:
            self.calls = 0

        def verify_split(self, request):  # noqa: ANN001
            self.calls += 1
            return GeminiVerificationResult(
                verdict="OK",
                issues=[],
                summary="ok",
            )

    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_cache.tmx"
    output_path = runtime_dir / "output_cache.tmx"
    cache_path = runtime_dir / "gemini-cache-test.json"
    _write_sample_tmx(input_path)

    verifier = CountingVerifier()
    repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        verify_with_gemini=True,
        gemini_verifier=verifier,
        enable_split_short_sentence_pair_guard=False,
        gemini_cache_path=cache_path,
    )
    assert verifier.calls == 1

    verifier_2 = CountingVerifier()
    repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        verify_with_gemini=True,
        gemini_verifier=verifier_2,
        enable_split_short_sentence_pair_guard=False,
        gemini_cache_path=cache_path,
    )
    assert verifier_2.calls == 0

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)
    cache_path.unlink(missing_ok=True)


def test_html_report_contains_interactive_tabs_for_cleanup_and_warnings():
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_tabs.tmx"
    output_path = runtime_dir / "output_tabs.tmx"
    html_report_path = runtime_dir / "report_tabs.html"
    input_path.write_text(
        """<?xml version="1.0" encoding="utf-8"?>
<tmx version="1.4">
  <header srclang="en-US" adminlang="en-US" creationtool="test" creationtoolversion="1.0" datatype="xml"/>
  <body>
    <tu creationid="u1">
      <tuv xml:lang="en-US"><seg>  Hello\u00A0 world  </seg></tuv>
      <tuv xml:lang="ru-RU"><seg>  Привет\u00A0 мир  </seg></tuv>
    </tu>
    <tu creationid="u2">
      <tuv xml:lang="en-US"><seg>Need translation now.</seg></tuv>
      <tuv xml:lang="ru-RU"><seg>!!!</seg></tuv>
    </tu>
  </body>
</tmx>
""",
        encoding="utf-8",
    )

    repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        html_report_path=html_report_path,
    )

    html_content = _read(html_report_path)
    assert "tab-button" in html_content
    assert "Split Changes" in html_content
    assert "Auto Cleanup" in html_content
    assert "Warnings" in html_content
    assert "Gemini Checks" in html_content
    assert "diff-add" in html_content or "diff-del" in html_content
    assert "Legend: changed regular spaces are marked as ·" in html_content
    assert "Whitespace delta:" in html_content

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)
    html_report_path.unlink(missing_ok=True)


def test_repair_can_disable_split_stage():
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_no_split.tmx"
    output_path = runtime_dir / "output_no_split.tmx"
    _write_sample_tmx(input_path)

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        enable_split=False,
    )

    assert stats.split_tus == 0
    assert stats.created_tus == 2
    content = _read(output_path)
    assert content.count("<tu ") == 2
    assert "Hello world. Next sentence!" in content

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)


def test_repair_matches_srclang_primary_language_tag():
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_srclang_primary.tmx"
    output_path = runtime_dir / "output_srclang_primary.tmx"
    input_path.write_text(
        """<?xml version="1.0" encoding="utf-8"?>
<tmx version="1.4">
  <header srclang="en" adminlang="en-US" creationtool="test" creationtoolversion="1.0" datatype="xml"/>
  <body>
    <tu creationid="u1">
      <tuv xml:lang="en-US"><seg>Hello world. Next sentence!</seg></tuv>
      <tuv xml:lang="ru-RU"><seg>Привет мир. Следующее предложение!</seg></tuv>
    </tu>
  </body>
</tmx>
""",
        encoding="utf-8",
    )

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        enable_split_short_sentence_pair_guard=False,
    )

    assert stats.split_tus == 1
    assert stats.created_tus == 2
    content = _read(output_path)
    assert content.count("<tu ") == 2

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)


def test_repair_falls_back_to_first_tuv_when_srclang_missing():
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_srclang_missing.tmx"
    output_path = runtime_dir / "output_srclang_missing.tmx"
    input_path.write_text(
        """<?xml version="1.0" encoding="utf-8"?>
<tmx version="1.4">
  <header srclang="" adminlang="en-US" creationtool="test" creationtoolversion="1.0" datatype="xml"/>
  <body>
    <tu creationid="u1">
      <tuv xml:lang="ru-RU"><seg>Привет мир. Следующее предложение!</seg></tuv>
      <tuv xml:lang="en-US"><seg>Hello world. Next sentence!</seg></tuv>
    </tu>
  </body>
</tmx>
""",
        encoding="utf-8",
    )

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        enable_split_short_sentence_pair_guard=False,
    )

    assert stats.split_tus == 1
    assert stats.created_tus == 2
    content = _read(output_path)
    assert content.count("<tu ") == 2

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)


def test_repair_can_remove_inline_tags_without_gluing_text():
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_tags_cleanup.tmx"
    output_path = runtime_dir / "output_tags_cleanup.tmx"
    input_path.write_text(
        """<?xml version="1.0" encoding="utf-8"?>
<tmx version="1.4">
  <header srclang="en-US" adminlang="en-US" creationtool="test" creationtoolversion="1.0" datatype="xml"/>
  <body>
    <tu creationid="u1">
      <tuv xml:lang="en-US"><seg>Hello!<ph x="1" type="0"/>World.</seg></tuv>
      <tuv xml:lang="ru-RU"><seg>Привет!<ph x="1" type="0"/>Мир.</seg></tuv>
    </tu>
  </body>
</tmx>
""",
        encoding="utf-8",
    )

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        enable_split=False,
        enable_cleanup_tag_removal=True,
    )

    assert stats.created_tus == 1
    content = _read(output_path)
    assert "<ph " not in content
    assert "Hello! World." in content
    assert "Привет! Мир." in content

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)


def test_repair_cleans_service_markup_in_context_content_prop():
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_context_cleanup.tmx"
    output_path = runtime_dir / "output_context_cleanup.tmx"
    input_path.write_text(
        """<?xml version="1.0" encoding="utf-8"?>
<tmx version="1.4">
  <header srclang="en-US" adminlang="en-US" creationtool="test" creationtoolversion="1.0" datatype="xml"/>
  <body>
    <tu creationid="u1">
      <prop type="x-ContextContent">Increases Health by ^{85 221 85}^%param1%%^{/color}^ and &lt;Color=#51D052FF&gt;%param2%%&lt;/Color&gt; with $m(s|s).</prop>
      <tuv xml:lang="en-US"><seg>Plain text.</seg></tuv>
      <tuv xml:lang="ru-RU"><seg>Обычный текст.</seg></tuv>
    </tu>
  </body>
</tmx>
""",
        encoding="utf-8",
    )

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        enable_split=False,
        enable_cleanup_percent_wrapped=True,
        enable_cleanup_game_markup=True,
    )

    assert stats.created_tus == 1
    content = _read(output_path)
    assert '<prop type="x-ContextContent">' in content
    assert "^{85 221 85}^" not in content
    assert "^{/color}^" not in content
    assert "&lt;Color=" not in content
    assert "&lt;/Color&gt;" not in content
    assert "%param1%" not in content
    assert "%param2%" not in content
    assert "$m(" not in content

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)


def test_html_cleanup_tab_shows_final_diff_and_intermediate_steps():
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_cleanup_aggregate.tmx"
    output_path = runtime_dir / "output_cleanup_aggregate.tmx"
    html_report_path = runtime_dir / "report_cleanup_aggregate.html"
    input_path.write_text(
        """<?xml version="1.0" encoding="utf-8"?>
<tmx version="1.4">
  <header srclang="en-US" adminlang="en-US" creationtool="test" creationtoolversion="1.0" datatype="xml"/>
  <body>
    <tu creationid="u1">
      <tuv xml:lang="en-US"><seg>Damage: ^{221 85 85}^%paramFloor%^{/color}^ ^{237 194 154}^(depends on the Totem power)^{/color}^</seg></tuv>
      <tuv xml:lang="ru-RU"><seg>Урон: ^{221 85 85}^%paramFloor%^{/color}^ ^{237 194 154}^(зависит от силы тотема)^{/color}^</seg></tuv>
    </tu>
  </body>
</tmx>
""",
        encoding="utf-8",
    )

    repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        dry_run=False,
        enable_split=False,
        enable_cleanup_percent_wrapped=True,
        enable_cleanup_game_markup=True,
        html_report_path=html_report_path,
    )

    html_content = _read(html_report_path)
    assert "Final cleanup result for TU (aggregated from 2 steps)." in html_content
    assert "Intermediate steps (2)" in html_content
    assert "depends on the Totem power" in html_content

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)
    html_report_path.unlink(missing_ok=True)


def test_html_report_limits_large_detail_sections():
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    html_report_path = runtime_dir / "report_limited.html"

    cleanup_events = [
        {
            "tu_index": idx,
            "scope": "segment",
            "rule": "normalize_spaces",
            "message": f"Cleanup #{idx}",
            "before_src": f"Before src {idx}",
            "after_src": f"After src {idx}",
            "before_tgt": f"Before tgt {idx}",
            "after_tgt": f"After tgt {idx}",
        }
        for idx in range(5)
    ]

    write_html_diff_report(
        path=html_report_path,
        input_path=Path("input.tmx"),
        output_path=Path("output.tmx"),
        stats=RepairStats(
            total_tus=5,
            split_tus=0,
            created_tus=5,
            src_lang="en-US",
            tgt_lang="ru-RU",
            skipped_tus=0,
            auto_actions=5,
        ),
        split_events=[],
        cleanup_events=cleanup_events,
        warning_events=[],
        gemini_audit_events=[],
        max_events_per_section=2,
    )

    html_content = _read(html_report_path)
    assert "Showing first 2 of 5 Auto Cleanup items. 3 omitted." in html_content
    assert "TU #1" in html_content
    assert "TU #2" in html_content
    assert "TU #3" not in html_content

    html_report_path.unlink(missing_ok=True)


def test_plan_mode_does_not_accumulate_report_detail_events(monkeypatch):
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_plan_report_details.tmx"
    output_path = runtime_dir / "output_plan_report_details.tmx"
    _write_multi_split_tmx(input_path, count=3)

    def fail_if_html_report_is_called(**kwargs):
        raise AssertionError("plan mode must not render HTML report details")

    monkeypatch.setattr("core.repair._write_html_diff_report", fail_if_html_report_is_called)

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        mode="plan",
        html_report_path=runtime_dir / "unused.html",
        enable_split_short_sentence_pair_guard=False,
    )

    assert stats.plan is not None
    assert len(stats.plan.proposals) == 3
    assert not output_path.exists()

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)


def test_plan_mode_disables_report_detail_collection_even_when_paths_are_known():
    assert not _should_collect_report_details(
        mode="plan",
        report_path=None,
        html_report_path=Path("known.html"),
        xlsx_report_path=Path("known.xlsx"),
        resume_state_path=None,
    )
    assert _should_collect_report_details(
        mode="apply",
        report_path=None,
        html_report_path=Path("known.html"),
        xlsx_report_path=None,
        resume_state_path=None,
    )


def test_apply_report_detail_events_are_bounded(monkeypatch):
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_bounded_details.tmx"
    output_path = runtime_dir / "output_bounded_details.tmx"
    report_path = runtime_dir / "report_bounded_details.json"
    html_report_path = runtime_dir / "report_bounded_details.html"

    body = "\n".join(
        f"""
    <tu creationid="u{idx}">
      <tuv xml:lang="en-US"><seg>This source sentence is intentionally much longer {idx}.</seg></tuv>
      <tuv xml:lang="ru-RU"><seg>да</seg></tuv>
    </tu>"""
        for idx in range(5)
    )
    input_path.write_text(
        f"""<?xml version="1.0" encoding="utf-8"?>
<tmx version="1.4">
  <header srclang="en-US" adminlang="en-US" creationtool="test" creationtoolversion="1.0" datatype="xml"/>
  <body>{body}
  </body>
</tmx>
""",
        encoding="utf-8",
    )
    monkeypatch.setattr("core.repair.MAX_REPORT_DETAIL_EVENTS_PER_KIND", 2)

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        report_path=report_path,
        html_report_path=html_report_path,
        enable_split=False,
    )

    assert stats.warn_issues == 5
    report = json.loads(_read(report_path))
    assert len(report["warning_events"]) == 2
    assert report["detail_event_limits"]["warnings"] == {
        "stored": 2,
        "total": 5,
        "omitted": 3,
    }
    html_content = _read(html_report_path)
    assert "Showing first 2 of 5 Warnings items. 3 omitted." in html_content

    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)
    report_path.unlink(missing_ok=True)
    html_report_path.unlink(missing_ok=True)


def test_plan_mode_compacts_proposal_details_after_limit(monkeypatch):
    runtime_dir = Path("tests") / "fixtures" / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    input_path = runtime_dir / "input_compact_plan.tmx"
    output_path = runtime_dir / "output_compact_plan.tmx"
    _write_multi_split_tmx(input_path, count=5)
    monkeypatch.setattr("core.repair.MAX_PLAN_DETAILED_PROPOSALS", 2)

    stats = repair_tmx_file(
        input_path=input_path,
        output_path=output_path,
        mode="plan",
        enable_split_short_sentence_pair_guard=False,
    )

    assert stats.plan is not None
    assert len(stats.plan.proposals) == 5
    assert stats.plan.proposals[0].src_parts
    assert stats.plan.proposals[1].src_parts
    assert stats.plan.proposals[2].proposal_id == "split:2"
    assert stats.plan.proposals[2].src_parts == []
    assert stats.plan.proposals[2].original_src == ""

    input_path.unlink(missing_ok=True)
