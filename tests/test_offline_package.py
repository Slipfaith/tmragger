"""Offline .tmrepair export/import workflow tests."""

from __future__ import annotations

import io
import json
from pathlib import Path
import zipfile

from core.offline_package import (
    export_tmrepair_package,
    import_tmrepair_package,
)
from core.plan import (
    Proposal,
    RepairPlan,
    make_cleanup_proposal_id,
    make_split_proposal_id,
)


SAMPLE_TMX = Path("sample") / "Eventum Premo_En-Ru.tmx"


def _sample_plan() -> RepairPlan:
    return RepairPlan(
        input_path=str(SAMPLE_TMX),
        total_tus=2,
        proposals=[
            Proposal(
                proposal_id=make_split_proposal_id(0),
                kind="split",
                tu_index=0,
                confidence="HIGH",
                src_parts=["One.", "Two."],
                tgt_parts=["Odin.", "Dva."],
                original_src="One. Two.",
                original_tgt="Odin. Dva.",
            ),
            Proposal(
                proposal_id=make_cleanup_proposal_id(1, "normalize_spaces", 0),
                kind="cleanup",
                tu_index=1,
                rule="normalize_spaces",
                message="Normalize spaces",
                before_src="A   B",
                after_src="A B",
                before_tgt="C   D",
                after_tgt="C D",
                original_src="A   B",
                original_tgt="C   D",
            ),
        ],
    )


def _inject_zip_file(path: Path, inner_path: str, payload: str) -> None:
    temp_path = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path, "r") as src, zipfile.ZipFile(
        temp_path, "w", compression=zipfile.ZIP_DEFLATED
    ) as dst:
        for info in src.infolist():
            dst.writestr(info, src.read(info.filename))
        dst.writestr(inner_path, payload)
    temp_path.replace(path)


def _inject_zip_bytes(path: Path, inner_path: str, payload: bytes) -> None:
    temp_path = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path, "r") as src, zipfile.ZipFile(
        temp_path, "w", compression=zipfile.ZIP_DEFLATED
    ) as dst:
        for info in src.infolist():
            if info.filename == inner_path:
                continue
            dst.writestr(info, src.read(info.filename))
        dst.writestr(inner_path, payload)
    temp_path.replace(path)


def _runtime_dir() -> Path:
    path = Path("tests") / "fixtures" / "runtime"
    path.mkdir(parents=True, exist_ok=True)
    return path


def test_export_tmrepair_package_creates_all_required_entries():
    assert SAMPLE_TMX.exists(), f"Missing sample TMX: {SAMPLE_TMX}"
    package_path = _runtime_dir() / "offline_export_sample.tmrepair"
    package_path.unlink(missing_ok=True)

    export_tmrepair_package(
        package_path=package_path,
        input_tmx_path=SAMPLE_TMX,
        plan=_sample_plan(),
        settings={
            "enable_split": True,
            "enable_cleanup_spaces": True,
            "enable_cleanup_service_markup": True,
            "enable_cleanup_garbage": True,
            "enable_cleanup_warnings": True,
            "enable_dedup_tus": False,
        },
    )

    assert package_path.exists()
    with zipfile.ZipFile(package_path, "r") as archive:
        names = set(archive.namelist())
        assert "manifest.json" in names
        assert "state.json" in names
        assert "source.tmx" in names
        assert "report.xlsx" in names
        # The HTML report was removed: XLSX is the single editing surface so
        # large plans no longer produce multi-MB single-page HTML documents.
        assert "report.html" not in names

        manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
        state = json.loads(archive.read("state.json").decode("utf-8"))
        report_xlsx = archive.read("report.xlsx")

    assert manifest["format_version"] == 1
    assert manifest["source_file_name"] == SAMPLE_TMX.name
    assert isinstance(manifest["source_sha256"], str)
    assert len(manifest["source_sha256"]) == 64
    assert len(state["issues"]) == 2
    assert state["issues"][0]["status"] == "pending"
    assert state["issues"][1]["status"] == "pending"

    from openpyxl import load_workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock

    workbook = load_workbook(io.BytesIO(report_xlsx), data_only=True, rich_text=True)
    try:
        sheet = workbook["Review"] if "Review" in workbook.sheetnames else workbook.active
        headers = [str(sheet.cell(row=4, column=col).value or "") for col in range(1, 8)]
        assert "Предлагаемая правка" not in headers
        assert "Source diff (- removed, + added)" not in headers
        assert "Target diff (- removed, + added)" not in headers
        source_col = headers.index("TU source") + 1
        target_col = headers.index("TU target") + 1
        decision_col = headers.index("Решение") + 1

        cleanup_row = None
        for row in range(5, sheet.max_row + 1):
            issue_id = str(sheet.cell(row=row, column=1).value or "")
            if issue_id == "cleanup:1:normalize_spaces:0":
                cleanup_row = row
                break
        assert cleanup_row is not None
        assert str(sheet.cell(row=cleanup_row, column=decision_col).value or "") == "pending"
        source_cell = sheet.cell(row=cleanup_row, column=source_col).value
        target_cell = sheet.cell(row=cleanup_row, column=target_col).value
        assert isinstance(source_cell, CellRichText)
        assert isinstance(target_cell, CellRichText)
        assert "·" in str(source_cell)
        assert "·" in str(target_cell)

        source_blocks = [part for part in source_cell if isinstance(part, TextBlock)]
        target_blocks = [part for part in target_cell if isinstance(part, TextBlock)]
        assert any(bool(block.font and block.font.strike) for block in source_blocks)
        assert any(bool(block.font and block.font.strike) for block in target_blocks)
    finally:
        workbook.close()
    package_path.unlink(missing_ok=True)


def test_import_tmrepair_package_prefers_decisions_json_and_updates_statuses():
    assert SAMPLE_TMX.exists(), f"Missing sample TMX: {SAMPLE_TMX}"
    package_path = _runtime_dir() / "offline_import_sample.tmrepair"
    package_path.unlink(missing_ok=True)

    export_tmrepair_package(
        package_path=package_path,
        input_tmx_path=SAMPLE_TMX,
        plan=_sample_plan(),
        settings={"enable_split": True},
    )

    decisions = {
        "decisions": [
            {"id": "split:0", "decision": "accept", "comment": "Looks good"},
            {"id": "cleanup:1:normalize_spaces:0", "decision": "skip"},
            {"id": "missing:999", "decision": "reject"},
        ]
    }
    _inject_zip_file(
        package_path,
        "decisions.json",
        json.dumps(decisions, ensure_ascii=False, indent=2),
    )

    result = import_tmrepair_package(package_path=package_path)
    assert result.accepted_count == 1
    assert result.rejected_count == 0
    assert result.skipped_count == 1
    assert result.unrecognized_count == 1

    accepted_split_ids = result.plan.accepted_split_ids()
    accepted_cleanup_ids = result.plan.accepted_cleanup_ids()
    assert accepted_split_ids == {"split:0"}
    assert accepted_cleanup_ids == set()
    result.source_tmx_path.unlink(missing_ok=True)
    package_path.unlink(missing_ok=True)


def test_import_reads_decisions_from_edited_xlsx():
    # The XLSX is now the only in-package editing surface, so decisions filled
    # into report.xlsx must round-trip back through import.
    from openpyxl import load_workbook

    assert SAMPLE_TMX.exists(), f"Missing sample TMX: {SAMPLE_TMX}"
    package_path = _runtime_dir() / "offline_xlsx_roundtrip.tmrepair"
    package_path.unlink(missing_ok=True)

    export_tmrepair_package(
        package_path=package_path,
        input_tmx_path=SAMPLE_TMX,
        plan=_sample_plan(),
        settings={"enable_split": True},
    )

    with zipfile.ZipFile(package_path, "r") as archive:
        xlsx_bytes = archive.read("report.xlsx")

    workbook = load_workbook(io.BytesIO(xlsx_bytes), rich_text=True)
    try:
        sheet = workbook["Review"] if "Review" in workbook.sheetnames else workbook.active
        header_row = None
        header_map: dict[str, int] = {}
        for row in range(1, min(30, sheet.max_row) + 1):
            values = [str(sheet.cell(row=row, column=col).value or "").strip() for col in range(1, 20)]
            if "ID проблемы" in values and "Решение" in values:
                header_row = row
                for col, value in enumerate(values, start=1):
                    if value:
                        header_map[value] = col
                break
        assert header_row is not None
        id_col = header_map["ID проблемы"]
        decision_col = header_map["Решение"]
        decisions_by_id = {"split:0": "accept", "cleanup:1:normalize_spaces:0": "reject"}
        for row in range(header_row + 1, sheet.max_row + 1):
            issue_id = str(sheet.cell(row=row, column=id_col).value or "").strip()
            if issue_id in decisions_by_id:
                sheet.cell(row=row, column=decision_col, value=decisions_by_id[issue_id])
        buffer = io.BytesIO()
        workbook.save(buffer)
    finally:
        workbook.close()

    _inject_zip_bytes(package_path, "report.xlsx", buffer.getvalue())

    result = import_tmrepair_package(package_path=package_path)
    assert result.decisions_source == "report.xlsx"
    assert result.accepted_count == 1
    assert result.rejected_count == 1
    assert result.plan.accepted_split_ids() == {"split:0"}
    assert result.plan.accepted_cleanup_ids() == set()

    result.source_tmx_path.unlink(missing_ok=True)
    package_path.unlink(missing_ok=True)
