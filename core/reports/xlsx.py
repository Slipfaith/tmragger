"""XLSX multi-sheet report (Summary / Splits / Cleanup / Warnings / Gemini).

``openpyxl`` is imported lazily so callers that never request an XLSX
report don't pay the import cost.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:  # pragma: no cover
    from core.repair import RepairStats


def write_xlsx_multi_sheet_report(
    path: Path,
    input_path: Path,
    output_path: Path,
    stats: "RepairStats",
    split_events: list[dict[str, object]],
    cleanup_events: list[dict[str, object]],
    warning_events: list[dict[str, object]],
    gemini_audit_events: list[dict[str, object]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_fill = PatternFill(start_color="DCEBFF", end_color="DCEBFF", fill_type="solid")
    header_font = Font(bold=True, color="0F172A")
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_rows: list[tuple[str, object]] = [
        ("Input TMX", str(input_path)),
        ("Output TMX", str(output_path)),
        ("Total TU", stats.total_tus),
        ("Split TU", stats.split_tus),
        ("Output TU", stats.created_tus),
        ("Skipped TU", stats.skipped_tus),
        ("High Confidence", stats.high_confidence_splits),
        ("Medium Confidence", stats.medium_confidence_splits),
        ("Gemini Checked", stats.gemini_checked),
        ("Gemini Rejected", stats.gemini_rejected),
        ("Gemini Tokens In", stats.gemini_input_tokens),
        ("Gemini Tokens Out", stats.gemini_output_tokens),
        ("Gemini Tokens Total", stats.gemini_total_tokens),
        ("Gemini Estimated Cost (USD)", f"{stats.gemini_estimated_cost_usd:.6f}"),
        ("AUTO Actions", stats.auto_actions),
        ("AUTO Removed TU", stats.auto_removed_tus),
        ("WARN Issues", stats.warn_issues),
    ]
    summary_ws.append(["Metric", "Value"])
    for key, value in summary_rows:
        summary_ws.append([key, value])
    _style_sheet(summary_ws, header_fill=header_fill, header_font=header_font, cell_alignment=cell_alignment)
    _autosize_columns(summary_ws)

    split_ws = wb.create_sheet("Split Changes")
    split_headers = [
        "TU #",
        "Confidence",
        "Gemini Verdict",
        "Source Before",
        "Target Before",
        "Source After Parts",
        "Target After Parts",
    ]
    split_ws.append(split_headers)
    for event in split_events:
        split_ws.append(
            [
                int(event.get("tu_index", 0)) + 1,
                str(event.get("confidence", "")),
                str(event.get("gemini_verdict", "") or ""),
                str(event.get("original_src", "")),
                str(event.get("original_tgt", "")),
                _join_parts_for_sheet(event.get("src_parts", [])),
                _join_parts_for_sheet(event.get("tgt_parts", [])),
            ]
        )
    _style_sheet(split_ws, header_fill=header_fill, header_font=header_font, cell_alignment=cell_alignment)
    _autosize_columns(split_ws)

    cleanup_ws = wb.create_sheet("Auto Cleanup")
    cleanup_headers = [
        "TU #",
        "Rule",
        "Message",
        "Source Before",
        "Source After",
        "Target Before",
        "Target After",
        "Remove Reason",
    ]
    cleanup_ws.append(cleanup_headers)
    for event in cleanup_events:
        cleanup_ws.append(
            [
                int(event.get("tu_index", 0)) + 1,
                str(event.get("rule", "")),
                str(event.get("message", "")),
                str(event.get("before_src", "")),
                str(event.get("after_src", "")),
                str(event.get("before_tgt", "")),
                str(event.get("after_tgt", "")),
                str(event.get("remove_reason", "") or ""),
            ]
        )
    _style_sheet(cleanup_ws, header_fill=header_fill, header_font=header_font, cell_alignment=cell_alignment)
    _autosize_columns(cleanup_ws)

    warnings_ws = wb.create_sheet("Warnings")
    warning_headers = [
        "TU #",
        "Rule",
        "Severity",
        "Message",
        "Source Snapshot",
        "Target Snapshot",
    ]
    warnings_ws.append(warning_headers)
    for event in warning_events:
        warnings_ws.append(
            [
                int(event.get("tu_index", 0)) + 1,
                str(event.get("rule", "")),
                str(event.get("severity", "")),
                str(event.get("message", "")),
                str(event.get("src_text", "")),
                str(event.get("tgt_text", "")),
            ]
        )
    _style_sheet(warnings_ws, header_fill=header_fill, header_font=header_font, cell_alignment=cell_alignment)
    _autosize_columns(warnings_ws)

    gemini_ws = wb.create_sheet("Gemini Checks")
    gemini_headers = [
        "TU #",
        "Kind",
        "Verdict",
        "Issues Count",
        "Summary",
        "Remove TU",
        "Remove Reason",
    ]
    gemini_ws.append(gemini_headers)
    for event in gemini_audit_events:
        gemini_ws.append(
            [
                int(event.get("tu_index", 0)) + 1,
                str(event.get("kind", "")),
                str(event.get("verdict", "")),
                int(event.get("issues_count", 0) or 0),
                str(event.get("summary", "")),
                str(event.get("remove_tu", "")),
                str(event.get("remove_reason", "") or ""),
            ]
        )
    _style_sheet(gemini_ws, header_fill=header_fill, header_font=header_font, cell_alignment=cell_alignment)
    _autosize_columns(gemini_ws)

    wb.save(path)


def _style_sheet(
    worksheet: object,
    *,
    header_fill: object,
    header_font: object,
    cell_alignment: object,
) -> None:
    rows = list(worksheet.iter_rows())  # type: ignore[attr-defined]
    if not rows:
        return
    for cell in rows[0]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows[1:]:
        for cell in row:
            cell.alignment = cell_alignment


def _autosize_columns(worksheet: object, max_width: int = 80) -> None:
    for column_cells in worksheet.columns:  # type: ignore[attr-defined]
        length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            length = max(length, len(str(value)))
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(  # type: ignore[attr-defined]
            max(12, length + 2), max_width
        )


def _join_parts_for_sheet(parts: object) -> str:
    if isinstance(parts, list):
        return "\n".join(f"{idx}. {str(part)}" for idx, part in enumerate(parts, start=1))
    return str(parts)
