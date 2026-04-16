"""TMX repair pipeline."""

from __future__ import annotations

from collections import Counter
from copy import deepcopy
from dataclasses import dataclass
from difflib import SequenceMatcher
from html import escape
import json
import logging
import os
from pathlib import Path
from typing import Callable
import xml.etree.ElementTree as ET

from core.events import (
    CleanupProposedEvent,
    FileCompleteEvent,
    FileStartEvent,
    GeminiResultEvent,
    GeminiUsage,
    RepairEvent,
    SplitProposedEvent,
    TuSkippedEvent,
    TuStartEvent,
    WarningEvent,
)
from core.gemini_client import (
    GeminiVerificationRequest,
    GeminiVerificationResult,
)
from core.gemini_prompt import GEMINI_CLEANUP_AUDIT_PROMPT
from core.plan import (
    Proposal,
    RepairPlan,
    make_cleanup_proposal_id,
    make_split_proposal_id,
)
from core.splitter import build_seg_from_inner_xml, propose_aligned_split, seg_to_inner_xml
from core.tm_cleanup import CleanupResult, analyze_and_clean_segments


XML_LANG = "{http://www.w3.org/XML/1998/namespace}lang"
ET.register_namespace("xml", "http://www.w3.org/XML/1998/namespace")

# <prop> types whose values reference neighbouring TUs (pre/post context, concordance,
# character offsets, etc.). After a split such props become misleading or outright wrong
# for every produced TU, so we drop them from the new split TUs. Matching is
# case-insensitive and covers both the canonical "x-Context*" SDL/Trados variants and a
# handful of common CAT-tool equivalents.
SPLIT_DROPPED_PROP_TYPES: frozenset[str] = frozenset(
    {
        "x-context",
        "x-contextpre",
        "x-contextpost",
        "x-contextcontent",
        "x-prev-segment",
        "x-next-segment",
        "x-previous-segment",
        "x-following-segment",
        "x-concordance",
        "x-sdl-contextpre",
        "x-sdl-contextpost",
    }
)


def _prop_type(elem: ET.Element) -> str:
    return (elem.attrib.get("type") or "").strip().lower()


def _should_drop_prop_in_split(elem: ET.Element) -> bool:
    if _local_name(elem.tag) != "prop":
        return False
    return _prop_type(elem) in SPLIT_DROPPED_PROP_TYPES


@dataclass
class RepairStats:
    total_tus: int
    split_tus: int
    created_tus: int
    src_lang: str
    tgt_lang: str | None
    skipped_tus: int
    high_confidence_splits: int = 0
    medium_confidence_splits: int = 0
    gemini_checked: int = 0
    gemini_rejected: int = 0
    gemini_input_tokens: int = 0
    gemini_output_tokens: int = 0
    gemini_total_tokens: int = 0
    gemini_estimated_cost_usd: float = 0.0
    auto_actions: int = 0
    auto_removed_tus: int = 0
    warn_issues: int = 0
    # Populated for every run; in plan mode the pipeline short-circuits after
    # building this. `None` only if the pipeline never reached the stats stage.
    plan: RepairPlan | None = None


DEFAULT_GEMINI_INPUT_PRICE_PER_1M_USD = 0.10
DEFAULT_GEMINI_OUTPUT_PRICE_PER_1M_USD = 0.40


def repair_tmx_file(
    input_path: Path,
    output_path: Path,
    dry_run: bool = False,
    logger: logging.Logger | None = None,
    verify_with_gemini: bool = False,
    gemini_verifier: object | None = None,
    max_gemini_checks: int | None = None,
    report_path: Path | None = None,
    gemini_prompt_template: str | None = None,
    html_report_path: Path | None = None,
    xlsx_report_path: Path | None = None,
    progress_callback: Callable[[dict[str, object]], None] | None = None,
    event_callback: Callable[[RepairEvent], None] | None = None,
    mode: str = "apply",
    accepted_split_ids: set[str] | None = None,
    accepted_cleanup_ids: set[str] | None = None,
) -> RepairStats:
    """Repair a TMX file by splitting aligned bilingual segments.

    Parameters
    ----------
    mode:
        ``"apply"`` (default) runs the full pipeline and writes the output TMX
        plus any configured reports. ``"plan"`` performs analysis only,
        populates ``RepairStats.plan`` with every candidate edit, and writes
        nothing — use this for a preview/approval UI. In plan mode the
        returned plan is serializable via :meth:`RepairPlan.to_json`.
    event_callback:
        Optional typed-event stream (see ``core.events``). Fired alongside the
        legacy ``progress_callback`` dict stream.
    accepted_split_ids / accepted_cleanup_ids:
        When provided, only proposals with matching IDs are applied. Proposals
        outside the set are skipped (for splits: TU is left untouched; for
        cleanup: the raw text remains as-is). Use ``None`` for "apply all".
    """
    log = logger or logging.getLogger("tmx_repair")
    plan_mode = mode == "plan"
    if mode not in {"apply", "plan"}:
        raise ValueError(f"Unknown mode: {mode!r}; expected 'apply' or 'plan'.")
    plan = RepairPlan(input_path=str(input_path))
    tree = ET.parse(input_path)
    root = tree.getroot()

    header = root.find("header")
    src_lang = header.attrib.get("srclang", "") if header is not None else ""

    body = root.find("body")
    if body is None:
        raise ValueError("TMX body is missing.")

    tus = list(body.findall("tu"))
    replacement_map: dict[int, list[ET.Element]] = {}
    split_tus = 0
    skipped_tus = 0
    tgt_lang_seen: str | None = None
    high_confidence_splits = 0
    medium_confidence_splits = 0
    gemini_checked = 0
    gemini_rejected = 0
    report_items: list[dict[str, object]] = []
    split_events: list[dict[str, object]] = []
    cleanup_events: list[dict[str, object]] = []
    warning_events: list[dict[str, object]] = []
    gemini_audit_events: list[dict[str, object]] = []
    active_prompt_template_for_run = gemini_prompt_template
    gemini_input_tokens = 0
    gemini_output_tokens = 0
    gemini_total_tokens = 0
    auto_actions_count = 0
    auto_removed_tus = 0
    warn_issues_count = 0
    input_price_per_1m = _read_env_float(
        "GEMINI_PRICE_INPUT_PER_1M_USD",
        DEFAULT_GEMINI_INPUT_PRICE_PER_1M_USD,
    )
    output_price_per_1m = _read_env_float(
        "GEMINI_PRICE_OUTPUT_PER_1M_USD",
        DEFAULT_GEMINI_OUTPUT_PRICE_PER_1M_USD,
    )

    plan.total_tus = len(tus)
    _emit_progress(
        progress_callback,
        {
            "event": "file_start",
            "input_path": str(input_path),
            "output_path": str(output_path),
            "total_tus": len(tus),
            "src_lang": src_lang,
            "verify_with_gemini": verify_with_gemini and gemini_verifier is not None,
            "gemini_input_tokens": 0,
            "gemini_output_tokens": 0,
            "gemini_total_tokens": 0,
            "gemini_estimated_cost_usd": 0.0,
        },
    )
    _emit_event(
        event_callback,
        FileStartEvent(
            input_path=str(input_path),
            total_tus=len(tus),
            src_lang=src_lang,
        ),
    )

    if verify_with_gemini and gemini_verifier is not None:
        active_template = active_prompt_template_for_run
        if active_template is None:
            active_template = getattr(gemini_verifier, "prompt_template", None)
        if not active_template:
            active_template = "<EMPTY_PROMPT_TEMPLATE>"
        active_prompt_template_for_run = active_template
        log.info("Gemini prompt template in use:\n%s", active_template)
        log.info(
            "Gemini estimated pricing: input=$%.4f/1M tokens, output=$%.4f/1M tokens",
            input_price_per_1m,
            output_price_per_1m,
        )

    total_tus = len(tus)
    for index, tu in enumerate(tus):
        tu_no = index + 1
        _emit_progress(
            progress_callback,
            {
                "event": "tu_start",
                "tu_index": tu_no,
                "total_tus": total_tus,
                "split_tus": split_tus,
                "skipped_tus": skipped_tus,
                "gemini_checked": gemini_checked,
                "gemini_rejected": gemini_rejected,
                "gemini_input_tokens": gemini_input_tokens,
                "gemini_output_tokens": gemini_output_tokens,
                "gemini_total_tokens": gemini_total_tokens,
                "gemini_estimated_cost_usd": _estimate_cost_usd(
                    gemini_input_tokens,
                    gemini_output_tokens,
                    input_price_per_1m,
                    output_price_per_1m,
                ),
            },
        )
        _emit_event(event_callback, TuStartEvent(tu_index=index, total_tus=total_tus))
        log.info("[TU %s/%s] Analyze started.", tu_no, total_tus)

        tuv_elements = [child for child in list(tu) if _local_name(child.tag) == "tuv"]
        if len(tuv_elements) < 2:
            skipped_tus += 1
            log.info("[TU %s/%s] Skip: less than 2 TUV entries.", tu_no, total_tus)
            _emit_progress(
                progress_callback,
                {
                    "event": "tu_skipped",
                    "tu_index": tu_no,
                    "total_tus": total_tus,
                    "reason": "less_than_two_tuv",
                    "split_tus": split_tus,
                    "skipped_tus": skipped_tus,
                    "gemini_checked": gemini_checked,
                    "gemini_rejected": gemini_rejected,
                    "gemini_input_tokens": gemini_input_tokens,
                    "gemini_output_tokens": gemini_output_tokens,
                    "gemini_total_tokens": gemini_total_tokens,
                    "gemini_estimated_cost_usd": _estimate_cost_usd(
                        gemini_input_tokens,
                        gemini_output_tokens,
                        input_price_per_1m,
                        output_price_per_1m,
                    ),
                },
            )
            continue

        if len(tuv_elements) > 2:
            tuv_langs = [_get_tuv_lang(t) for t in tuv_elements]
            skipped_tus += 1
            warning_events.append(
                {
                    "tu_index": index,
                    "tu_no": tu_no,
                    "rule": "multilang_tu_skipped",
                    "severity": "WARN",
                    "message": (
                        "TU has more than 2 <tuv> entries "
                        f"(languages: {', '.join(tuv_langs)}); left unchanged."
                    ),
                    "src_text": "",
                    "tgt_text": "",
                    "details": {"langs": tuv_langs, "count": len(tuv_elements)},
                }
            )
            warn_issues_count += 1
            log.info(
                "[TU %s/%s] Skip: multi-language TU with %s TUVs (%s).",
                tu_no,
                total_tus,
                len(tuv_elements),
                ", ".join(tuv_langs),
            )
            _emit_progress(
                progress_callback,
                {
                    "event": "tu_skipped",
                    "tu_index": tu_no,
                    "total_tus": total_tus,
                    "reason": "multilang_tu",
                    "tuv_count": len(tuv_elements),
                    "tuv_langs": tuv_langs,
                    "split_tus": split_tus,
                    "skipped_tus": skipped_tus,
                    "gemini_checked": gemini_checked,
                    "gemini_rejected": gemini_rejected,
                    "gemini_input_tokens": gemini_input_tokens,
                    "gemini_output_tokens": gemini_output_tokens,
                    "gemini_total_tokens": gemini_total_tokens,
                    "gemini_estimated_cost_usd": _estimate_cost_usd(
                        gemini_input_tokens,
                        gemini_output_tokens,
                        input_price_per_1m,
                        output_price_per_1m,
                    ),
                },
            )
            continue

        src_tuv = _find_tuv_by_lang(tuv_elements, src_lang)
        if src_tuv is None:
            skipped_tus += 1
            log.info("[TU %s/%s] Skip: source TUV for '%s' not found.", tu_no, total_tus, src_lang)
            _emit_progress(
                progress_callback,
                {
                    "event": "tu_skipped",
                    "tu_index": tu_no,
                    "total_tus": total_tus,
                    "reason": "missing_src_tuv",
                    "split_tus": split_tus,
                    "skipped_tus": skipped_tus,
                    "gemini_checked": gemini_checked,
                    "gemini_rejected": gemini_rejected,
                    "gemini_input_tokens": gemini_input_tokens,
                    "gemini_output_tokens": gemini_output_tokens,
                    "gemini_total_tokens": gemini_total_tokens,
                    "gemini_estimated_cost_usd": _estimate_cost_usd(
                        gemini_input_tokens,
                        gemini_output_tokens,
                        input_price_per_1m,
                        output_price_per_1m,
                    ),
                },
            )
            continue

        tgt_tuv = next((t for t in tuv_elements if _get_tuv_lang(t) != src_lang), None)
        if tgt_tuv is None:
            skipped_tus += 1
            log.info("[TU %s/%s] Skip: target TUV not found.", tu_no, total_tus)
            _emit_progress(
                progress_callback,
                {
                    "event": "tu_skipped",
                    "tu_index": tu_no,
                    "total_tus": total_tus,
                    "reason": "missing_tgt_tuv",
                    "split_tus": split_tus,
                    "skipped_tus": skipped_tus,
                    "gemini_checked": gemini_checked,
                    "gemini_rejected": gemini_rejected,
                    "gemini_input_tokens": gemini_input_tokens,
                    "gemini_output_tokens": gemini_output_tokens,
                    "gemini_total_tokens": gemini_total_tokens,
                    "gemini_estimated_cost_usd": _estimate_cost_usd(
                        gemini_input_tokens,
                        gemini_output_tokens,
                        input_price_per_1m,
                        output_price_per_1m,
                    ),
                },
            )
            continue

        tgt_lang_seen = _get_tuv_lang(tgt_tuv)
        src_seg = src_tuv.find("seg")
        tgt_seg = tgt_tuv.find("seg")
        if src_seg is None or tgt_seg is None:
            skipped_tus += 1
            log.info("[TU %s/%s] Skip: source or target SEG is missing.", tu_no, total_tus)
            _emit_progress(
                progress_callback,
                {
                    "event": "tu_skipped",
                    "tu_index": tu_no,
                    "total_tus": total_tus,
                    "reason": "missing_seg",
                    "split_tus": split_tus,
                    "skipped_tus": skipped_tus,
                    "gemini_checked": gemini_checked,
                    "gemini_rejected": gemini_rejected,
                    "gemini_input_tokens": gemini_input_tokens,
                    "gemini_output_tokens": gemini_output_tokens,
                    "gemini_total_tokens": gemini_total_tokens,
                    "gemini_estimated_cost_usd": _estimate_cost_usd(
                        gemini_input_tokens,
                        gemini_output_tokens,
                        input_price_per_1m,
                        output_price_per_1m,
                    ),
                },
            )
            continue

        original_src_text = seg_to_inner_xml(src_seg)
        original_tgt_text = seg_to_inner_xml(tgt_seg)
        cleanup_result = analyze_and_clean_segments(
            src_inner_xml=original_src_text,
            tgt_inner_xml=original_tgt_text,
            src_lang=src_lang,
            tgt_lang=tgt_lang_seen,
        )
        auto_actions_count += len(cleanup_result.auto_actions)
        warn_issues_count += len(cleanup_result.warnings)

        for ordinal, action in enumerate(cleanup_result.auto_actions):
            rule_name = str(action.get("rule", ""))
            action_event = {
                "tu_index": index,
                "tu_no": tu_no,
                "rule": rule_name,
                "message": action.get("message", ""),
                "before_src": action.get("before_src", original_src_text),
                "after_src": action.get("after_src", cleanup_result.src_inner_xml),
                "before_tgt": action.get("before_tgt", original_tgt_text),
                "after_tgt": action.get("after_tgt", cleanup_result.tgt_inner_xml),
                "remove_reason": action.get("remove_reason"),
            }
            cleanup_events.append(action_event)
            cleanup_proposal_id = make_cleanup_proposal_id(index, rule_name, ordinal)
            plan.proposals.append(
                Proposal(
                    proposal_id=cleanup_proposal_id,
                    kind="cleanup",
                    tu_index=index,
                    rule=rule_name,
                    message=str(action.get("message", "")),
                    before_src=str(action_event["before_src"]),
                    after_src=str(action_event["after_src"]),
                    before_tgt=str(action_event["before_tgt"]),
                    after_tgt=str(action_event["after_tgt"]),
                    original_src=original_src_text,
                    original_tgt=original_tgt_text,
                )
            )
            _emit_event(
                event_callback,
                CleanupProposedEvent(
                    tu_index=index,
                    rule=rule_name,
                    message=str(action.get("message", "")),
                    before_src=str(action_event["before_src"]),
                    after_src=str(action_event["after_src"]),
                    before_tgt=str(action_event["before_tgt"]),
                    after_tgt=str(action_event["after_tgt"]),
                ),
            )
            log.info(
                "[TU %s/%s] AUTO %s: %s",
                tu_no,
                total_tus,
                action_event["rule"],
                action_event["message"],
            )

        for warning in cleanup_result.warnings:
            warning_event = {
                "tu_index": index,
                "tu_no": tu_no,
                "rule": warning.get("rule", ""),
                "severity": warning.get("severity", "WARN"),
                "message": warning.get("message", ""),
                "src_text": cleanup_result.src_plain_text,
                "tgt_text": cleanup_result.tgt_plain_text,
                "details": warning,
            }
            warning_events.append(warning_event)
            _emit_event(
                event_callback,
                WarningEvent(
                    tu_index=index,
                    rule=str(warning.get("rule", "")),
                    severity=str(warning.get("severity", "WARN")),
                    message=str(warning.get("message", "")),
                ),
            )
            log.info(
                "[TU %s/%s] WARN %s: %s",
                tu_no,
                total_tus,
                warning_event["rule"],
                warning_event["message"],
            )

        if cleanup_result.src_inner_xml != original_src_text:
            _set_seg_inner_xml(src_seg, cleanup_result.src_inner_xml)
        if cleanup_result.tgt_inner_xml != original_tgt_text:
            _set_seg_inner_xml(tgt_seg, cleanup_result.tgt_inner_xml)

        cleaned_src_text = cleanup_result.src_inner_xml
        cleaned_tgt_text = cleanup_result.tgt_inner_xml

        cleanup_gemini_result = _verify_cleanup_with_gemini(
            verify_with_gemini=verify_with_gemini,
            gemini_verifier=gemini_verifier,
            src_lang=src_lang,
            tgt_lang=tgt_lang_seen,
            tu_no=tu_no,
            total_tus=total_tus,
            original_src_text=original_src_text,
            original_tgt_text=original_tgt_text,
            cleaned_src_text=cleaned_src_text,
            cleaned_tgt_text=cleaned_tgt_text,
            cleanup_result=cleanup_result,
            input_price_per_1m=input_price_per_1m,
            output_price_per_1m=output_price_per_1m,
            gemini_checked=gemini_checked,
            gemini_rejected=gemini_rejected,
            gemini_input_tokens=gemini_input_tokens,
            gemini_output_tokens=gemini_output_tokens,
            gemini_total_tokens=gemini_total_tokens,
            progress_callback=progress_callback,
            report_items=report_items,
            gemini_audit_events=gemini_audit_events,
            log=log,
        )
        if cleanup_gemini_result is not None:
            gemini_checked = cleanup_gemini_result["gemini_checked"]
            gemini_rejected = cleanup_gemini_result["gemini_rejected"]
            gemini_input_tokens = cleanup_gemini_result["gemini_input_tokens"]
            gemini_output_tokens = cleanup_gemini_result["gemini_output_tokens"]
            gemini_total_tokens = cleanup_gemini_result["gemini_total_tokens"]
            if cleanup_result.remove_tu and cleanup_gemini_result["verdict"] == "FAIL":
                cleanup_result.remove_tu = False
                cleanup_result.remove_reason = None
                warning_events.append(
                    {
                        "tu_index": index,
                        "tu_no": tu_no,
                        "rule": "cleanup_remove_reverted_by_gemini",
                        "severity": "WARN",
                        "message": "Gemini flagged cleanup removal; TU kept in output.",
                        "src_text": cleanup_result.src_plain_text,
                        "tgt_text": cleanup_result.tgt_plain_text,
                        "details": {},
                    }
                )
                warn_issues_count += 1
                log.info("[TU %s/%s] Cleanup removal reverted by Gemini.", tu_no, total_tus)

        if cleanup_result.remove_tu:
            auto_removed_tus += 1
            skipped_tus += 1
            replacement_map[index] = []
            log.info(
                "[TU %s/%s] AUTO removed as garbage: %s",
                tu_no,
                total_tus,
                cleanup_result.remove_reason or "unknown",
            )
            _emit_progress(
                progress_callback,
                {
                    "event": "tu_removed_cleanup",
                    "tu_index": tu_no,
                    "total_tus": total_tus,
                    "reason": cleanup_result.remove_reason or "cleanup_remove",
                    "split_tus": split_tus,
                    "skipped_tus": skipped_tus,
                    "gemini_checked": gemini_checked,
                    "gemini_rejected": gemini_rejected,
                    "gemini_input_tokens": gemini_input_tokens,
                    "gemini_output_tokens": gemini_output_tokens,
                    "gemini_total_tokens": gemini_total_tokens,
                    "gemini_estimated_cost_usd": _estimate_cost_usd(
                        gemini_input_tokens,
                        gemini_output_tokens,
                        input_price_per_1m,
                        output_price_per_1m,
                    ),
                },
            )
            continue

        log.info(
            "[TU %s/%s] Candidate text | src=%s | tgt=%s",
            tu_no,
            total_tus,
            _preview(cleaned_src_text),
            _preview(cleaned_tgt_text),
        )
        proposal = propose_aligned_split(cleaned_src_text, cleaned_tgt_text)
        if proposal is None:
            skipped_tus += 1
            log.info("[TU %s/%s] Skip: no aligned split proposal.", tu_no, total_tus)
            _emit_progress(
                progress_callback,
                {
                    "event": "tu_skipped",
                    "tu_index": tu_no,
                    "total_tus": total_tus,
                    "reason": "no_split_proposal",
                    "split_tus": split_tus,
                    "skipped_tus": skipped_tus,
                    "gemini_checked": gemini_checked,
                    "gemini_rejected": gemini_rejected,
                    "gemini_input_tokens": gemini_input_tokens,
                    "gemini_output_tokens": gemini_output_tokens,
                    "gemini_total_tokens": gemini_total_tokens,
                    "gemini_estimated_cost_usd": _estimate_cost_usd(
                        gemini_input_tokens,
                        gemini_output_tokens,
                        input_price_per_1m,
                        output_price_per_1m,
                    ),
                },
            )
            continue

        src_parts, tgt_parts = proposal
        if len(src_parts) < 2:
            skipped_tus += 1
            log.info("[TU %s/%s] Skip: proposal produced < 2 parts.", tu_no, total_tus)
            _emit_progress(
                progress_callback,
                {
                    "event": "tu_skipped",
                    "tu_index": tu_no,
                    "total_tus": total_tus,
                    "reason": "proposal_less_than_two_parts",
                    "split_tus": split_tus,
                    "skipped_tus": skipped_tus,
                    "gemini_checked": gemini_checked,
                    "gemini_rejected": gemini_rejected,
                    "gemini_input_tokens": gemini_input_tokens,
                    "gemini_output_tokens": gemini_output_tokens,
                    "gemini_total_tokens": gemini_total_tokens,
                    "gemini_estimated_cost_usd": _estimate_cost_usd(
                        gemini_input_tokens,
                        gemini_output_tokens,
                        input_price_per_1m,
                        output_price_per_1m,
                    ),
                },
            )
            continue

        log.info("[TU %s/%s] Split proposal found: %s parts.", tu_no, total_tus, len(src_parts))
        for pair_index, (src_part, tgt_part) in enumerate(zip(src_parts, tgt_parts), start=1):
            log.info(
                "[TU %s/%s] Pair %s | src=%s | tgt=%s",
                tu_no,
                total_tus,
                pair_index,
                _preview(src_part),
                _preview(tgt_part),
            )

        confidence = "HIGH"
        gemini_result: GeminiVerificationResult | None = None
        can_verify = (
            verify_with_gemini
            and gemini_verifier is not None
            and (max_gemini_checks is None or gemini_checked < max_gemini_checks)
        )
        if can_verify:
            gemini_checked += 1
            log.info("[TU %s/%s] Gemini verification started.", tu_no, total_tus)
            verify_request = GeminiVerificationRequest(
                src_lang=src_lang,
                tgt_lang=tgt_lang_seen,
                original_src=cleaned_src_text,
                original_tgt=cleaned_tgt_text,
                src_parts=src_parts,
                tgt_parts=tgt_parts,
            )
            if gemini_prompt_template is not None:
                gemini_result = _run_gemini_verification(
                    gemini_verifier=gemini_verifier,
                    verify_request=verify_request,
                    prompt_template=gemini_prompt_template,
                )
            else:
                gemini_result = _run_gemini_verification(
                    gemini_verifier=gemini_verifier,
                    verify_request=verify_request,
                )

            gemini_input_tokens += max(0, int(gemini_result.prompt_tokens))
            gemini_output_tokens += max(0, int(gemini_result.completion_tokens))
            result_total_tokens = max(
                0,
                int(gemini_result.total_tokens),
            )
            if result_total_tokens == 0:
                result_total_tokens = max(0, int(gemini_result.prompt_tokens)) + max(
                    0,
                    int(gemini_result.completion_tokens),
                )
            gemini_total_tokens += result_total_tokens
            run_cost = _estimate_cost_usd(
                gemini_input_tokens,
                gemini_output_tokens,
                input_price_per_1m,
                output_price_per_1m,
            )

            log.info(
                (
                    "[TU %s/%s] Gemini verdict=%s issues=%s summary=%s "
                    "tokens(in=%s out=%s total=%s) run_tokens(in=%s out=%s total=%s) est_cost=$%.6f"
                ),
                tu_no,
                total_tus,
                gemini_result.verdict,
                len(gemini_result.issues),
                gemini_result.summary,
                gemini_result.prompt_tokens,
                gemini_result.completion_tokens,
                result_total_tokens,
                gemini_input_tokens,
                gemini_output_tokens,
                gemini_total_tokens,
                run_cost,
            )
            _emit_progress(
                progress_callback,
                {
                    "event": "gemini_result",
                    "tu_index": tu_no,
                    "total_tus": total_tus,
                    "verdict": gemini_result.verdict,
                    "summary": gemini_result.summary,
                    "issues_count": len(gemini_result.issues),
                    "split_tus": split_tus,
                    "skipped_tus": skipped_tus,
                    "gemini_checked": gemini_checked,
                    "gemini_rejected": gemini_rejected,
                    "gemini_input_tokens": gemini_input_tokens,
                    "gemini_output_tokens": gemini_output_tokens,
                    "gemini_total_tokens": gemini_total_tokens,
                    "gemini_estimated_cost_usd": run_cost,
                },
            )
            report_items.append(
                {
                    "category": "split_verification",
                    "tu_index": index,
                    "verdict": gemini_result.verdict,
                    "summary": gemini_result.summary,
                    "issues": [issue.__dict__ for issue in gemini_result.issues],
                    "prompt_tokens": gemini_result.prompt_tokens,
                    "completion_tokens": gemini_result.completion_tokens,
                    "total_tokens": result_total_tokens,
                    "run_gemini_input_tokens": gemini_input_tokens,
                    "run_gemini_output_tokens": gemini_output_tokens,
                    "run_gemini_total_tokens": gemini_total_tokens,
                    "run_estimated_cost_usd": run_cost,
                }
            )
            gemini_audit_events.append(
                {
                    "tu_index": index,
                    "kind": "split",
                    "verdict": gemini_result.verdict,
                    "summary": gemini_result.summary,
                    "issues_count": len(gemini_result.issues),
                }
            )
            if gemini_result.verdict == "FAIL":
                gemini_rejected += 1
                skipped_tus += 1
                log.info("[TU %s/%s] Split rejected by Gemini. Keeping original TU.", tu_no, total_tus)
                _emit_progress(
                    progress_callback,
                    {
                        "event": "tu_rejected",
                        "tu_index": tu_no,
                        "total_tus": total_tus,
                        "reason": "gemini_fail",
                        "split_tus": split_tus,
                        "skipped_tus": skipped_tus,
                        "gemini_checked": gemini_checked,
                        "gemini_rejected": gemini_rejected,
                        "gemini_input_tokens": gemini_input_tokens,
                        "gemini_output_tokens": gemini_output_tokens,
                        "gemini_total_tokens": gemini_total_tokens,
                        "gemini_estimated_cost_usd": run_cost,
                    },
                )
                continue
            confidence = "MEDIUM"

        split_proposal_id = make_split_proposal_id(index)
        split_accepted_by_user = (
            accepted_split_ids is None or split_proposal_id in accepted_split_ids
        )
        plan.proposals.append(
            Proposal(
                proposal_id=split_proposal_id,
                kind="split",
                tu_index=index,
                accepted=split_accepted_by_user,
                confidence=confidence,
                src_parts=list(src_parts),
                tgt_parts=list(tgt_parts),
                original_src=cleaned_src_text,
                original_tgt=cleaned_tgt_text,
            )
        )
        _emit_event(
            event_callback,
            SplitProposedEvent(
                tu_index=index,
                src_parts=list(src_parts),
                tgt_parts=list(tgt_parts),
                confidence=confidence,  # type: ignore[arg-type]
                original_src=cleaned_src_text,
                original_tgt=cleaned_tgt_text,
            ),
        )

        if not split_accepted_by_user:
            skipped_tus += 1
            log.info(
                "[TU %s/%s] Split rejected by caller (accepted_split_ids filter).",
                tu_no,
                total_tus,
            )
            _emit_progress(
                progress_callback,
                {
                    "event": "tu_rejected",
                    "tu_index": tu_no,
                    "total_tus": total_tus,
                    "reason": "user_rejected",
                    "split_tus": split_tus,
                    "skipped_tus": skipped_tus,
                    "gemini_checked": gemini_checked,
                    "gemini_rejected": gemini_rejected,
                    "gemini_input_tokens": gemini_input_tokens,
                    "gemini_output_tokens": gemini_output_tokens,
                    "gemini_total_tokens": gemini_total_tokens,
                    "gemini_estimated_cost_usd": _estimate_cost_usd(
                        gemini_input_tokens,
                        gemini_output_tokens,
                        input_price_per_1m,
                        output_price_per_1m,
                    ),
                },
            )
            continue

        replacement_map[index] = _build_split_tus(
            tu=tu,
            src_lang=src_lang,
            tgt_lang=tgt_lang_seen,
            src_parts=src_parts,
            tgt_parts=tgt_parts,
            confidence=confidence,
            gemini_result=gemini_result,
        )
        split_tus += 1
        split_events.append(
            {
                "tu_index": index,
                "confidence": confidence,
                "gemini_verdict": gemini_result.verdict if gemini_result is not None else None,
                "original_src": cleaned_src_text,
                "original_tgt": cleaned_tgt_text,
                "src_parts": src_parts,
                "tgt_parts": tgt_parts,
            }
        )
        if confidence == "HIGH":
            high_confidence_splits += 1
        else:
            medium_confidence_splits += 1
        log.info(
            "[TU %s/%s] Split accepted. confidence=%s, output_parts=%s",
            tu_no,
            total_tus,
            confidence,
            len(src_parts),
        )
        _emit_progress(
            progress_callback,
            {
                "event": "tu_split_applied",
                "tu_index": tu_no,
                "total_tus": total_tus,
                "confidence": confidence,
                "split_tus": split_tus,
                "skipped_tus": skipped_tus,
                "gemini_checked": gemini_checked,
                "gemini_rejected": gemini_rejected,
                "gemini_input_tokens": gemini_input_tokens,
                "gemini_output_tokens": gemini_output_tokens,
                "gemini_total_tokens": gemini_total_tokens,
                "gemini_estimated_cost_usd": _estimate_cost_usd(
                    gemini_input_tokens,
                    gemini_output_tokens,
                    input_price_per_1m,
                    output_price_per_1m,
                ),
            },
        )

    if replacement_map:
        body.clear()
        for index, tu in enumerate(tus):
            replacements = replacement_map.get(index)
            if replacements is not None:
                body.extend(replacements)
            else:
                body.append(tu)

    created_tus = len(list(body.findall("tu")))
    gemini_estimated_cost_usd = _estimate_cost_usd(
        gemini_input_tokens,
        gemini_output_tokens,
        input_price_per_1m,
        output_price_per_1m,
    )
    stats = RepairStats(
        total_tus=len(tus),
        split_tus=split_tus,
        created_tus=created_tus,
        src_lang=src_lang,
        tgt_lang=tgt_lang_seen,
        skipped_tus=skipped_tus,
        high_confidence_splits=high_confidence_splits,
        medium_confidence_splits=medium_confidence_splits,
        gemini_checked=gemini_checked,
        gemini_rejected=gemini_rejected,
        gemini_input_tokens=gemini_input_tokens,
        gemini_output_tokens=gemini_output_tokens,
        gemini_total_tokens=gemini_total_tokens,
        gemini_estimated_cost_usd=gemini_estimated_cost_usd,
        auto_actions=auto_actions_count,
        auto_removed_tus=auto_removed_tus,
        warn_issues=warn_issues_count,
    )
    log.info(
        (
            "TMX processed: total=%s, split=%s, skipped=%s, output_tu=%s, "
            "high=%s, medium=%s, gemini_checked=%s, gemini_rejected=%s, "
            "gemini_tokens_in=%s, gemini_tokens_out=%s, gemini_tokens_total=%s, est_cost=$%.6f, "
            "auto_actions=%s, auto_removed_tus=%s, warn_issues=%s"
        ),
        stats.total_tus,
        stats.split_tus,
        stats.skipped_tus,
        stats.created_tus,
        stats.high_confidence_splits,
        stats.medium_confidence_splits,
        stats.gemini_checked,
        stats.gemini_rejected,
        stats.gemini_input_tokens,
        stats.gemini_output_tokens,
        stats.gemini_total_tokens,
        stats.gemini_estimated_cost_usd,
        stats.auto_actions,
        stats.auto_removed_tus,
        stats.warn_issues,
    )
    stats.plan = plan
    _emit_progress(
        progress_callback,
        {
            "event": "file_complete",
            "input_path": str(input_path),
            "output_path": str(output_path),
            "total_tus": stats.total_tus,
            "split_tus": stats.split_tus,
            "skipped_tus": stats.skipped_tus,
            "output_tu": stats.created_tus,
            "high_confidence_splits": stats.high_confidence_splits,
            "medium_confidence_splits": stats.medium_confidence_splits,
            "gemini_checked": stats.gemini_checked,
            "gemini_rejected": stats.gemini_rejected,
            "gemini_input_tokens": stats.gemini_input_tokens,
            "gemini_output_tokens": stats.gemini_output_tokens,
            "gemini_total_tokens": stats.gemini_total_tokens,
            "gemini_estimated_cost_usd": stats.gemini_estimated_cost_usd,
            "auto_actions": stats.auto_actions,
            "auto_removed_tus": stats.auto_removed_tus,
            "warn_issues": stats.warn_issues,
        },
    )
    _emit_event(
        event_callback,
        FileCompleteEvent(
            input_path=str(input_path),
            output_path=str(output_path),
            total_tus=stats.total_tus,
            split_tus=stats.split_tus,
            created_tus=stats.created_tus,
            skipped_tus=stats.skipped_tus,
        ),
    )

    if plan_mode:
        log.info("Plan mode: no output or reports written; %d proposals collected.", len(plan.proposals))
        return stats

    if not dry_run:
        tree.write(output_path, encoding="utf-8", xml_declaration=True, short_empty_elements=False)
        log.info("Saved repaired TMX to %s", output_path)
    else:
        log.info("Dry run enabled. Output file was not written.")

    if report_path is not None:
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report = {
            "input_path": str(input_path),
            "output_path": str(output_path),
            "total_tus": stats.total_tus,
            "split_tus": stats.split_tus,
            "created_tus": stats.created_tus,
            "skipped_tus": stats.skipped_tus,
            "high_confidence_splits": stats.high_confidence_splits,
            "medium_confidence_splits": stats.medium_confidence_splits,
            "gemini_checked": stats.gemini_checked,
            "gemini_rejected": stats.gemini_rejected,
            "gemini_input_tokens": stats.gemini_input_tokens,
            "gemini_output_tokens": stats.gemini_output_tokens,
            "gemini_total_tokens": stats.gemini_total_tokens,
            "gemini_estimated_cost_usd": stats.gemini_estimated_cost_usd,
            "gemini_pricing_input_per_1m_usd": input_price_per_1m,
            "gemini_pricing_output_per_1m_usd": output_price_per_1m,
            "gemini_prompt_template": active_prompt_template_for_run,
            "gemini_cleanup_prompt_template": GEMINI_CLEANUP_AUDIT_PROMPT,
            "auto_actions": stats.auto_actions,
            "auto_removed_tus": stats.auto_removed_tus,
            "warn_issues": stats.warn_issues,
            "cleanup_events": cleanup_events,
            "warning_events": warning_events,
            "gemini_audit_events": gemini_audit_events,
            "items": report_items,
        }
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        log.info("Verification report saved to %s", report_path)

    if html_report_path is not None:
        html_report_path.parent.mkdir(parents=True, exist_ok=True)
        _write_html_diff_report(
            path=html_report_path,
            input_path=input_path,
            output_path=output_path,
            stats=stats,
            split_events=split_events,
            cleanup_events=cleanup_events,
            warning_events=warning_events,
            gemini_audit_events=gemini_audit_events,
        )
        log.info("HTML diff report saved to %s", html_report_path)

    if xlsx_report_path is not None:
        xlsx_report_path.parent.mkdir(parents=True, exist_ok=True)
        _write_xlsx_multi_sheet_report(
            path=xlsx_report_path,
            input_path=input_path,
            output_path=output_path,
            stats=stats,
            split_events=split_events,
            cleanup_events=cleanup_events,
            warning_events=warning_events,
            gemini_audit_events=gemini_audit_events,
        )
        log.info("XLSX multi-sheet report saved to %s", xlsx_report_path)

    return stats


def _emit_progress(
    progress_callback: Callable[[dict[str, object]], None] | None,
    payload: dict[str, object],
) -> None:
    if progress_callback is None:
        return
    try:
        progress_callback(payload)
    except Exception:
        # Progress callbacks are optional and must never break the repair run.
        return


def _emit_event(
    event_callback: Callable[[RepairEvent], None] | None,
    event: RepairEvent,
) -> None:
    if event_callback is None:
        return
    try:
        event_callback(event)
    except Exception:
        # Typed-event listeners are optional and must never break the run.
        return


def _read_env_float(name: str, default: float) -> float:
    raw = os.getenv(name)
    if raw is None:
        return default
    try:
        value = float(raw.strip())
    except ValueError:
        return default
    return value if value >= 0 else default


def _estimate_cost_usd(
    input_tokens: int,
    output_tokens: int,
    input_price_per_1m: float,
    output_price_per_1m: float,
) -> float:
    return (max(0, input_tokens) / 1_000_000.0) * input_price_per_1m + (
        max(0, output_tokens) / 1_000_000.0
    ) * output_price_per_1m


def _preview(text: str, limit: int = 180) -> str:
    compact = " ".join(text.replace("\n", " ").split())
    if len(compact) <= limit:
        return compact
    return f"{compact[: max(1, limit - 3)]}..."


def _render_inline_diff(before: str, after: str) -> str:
    matcher = SequenceMatcher(a=before, b=after, autojunk=False)
    before_chunks: list[str] = []
    after_chunks: list[str] = []
    for op, i1, i2, j1, j2 in matcher.get_opcodes():
        before_part = before[i1:i2]
        after_part = after[j1:j2]
        if op == "equal":
            if before_part:
                before_chunks.append(f'<span class="diff-eq">{_visible_text(before_part, mark_spaces=False)}</span>')
            if after_part:
                after_chunks.append(f'<span class="diff-eq">{_visible_text(after_part, mark_spaces=False)}</span>')
        elif op == "delete":
            if before_part:
                before_chunks.append(f'<span class="diff-del">{_visible_text(before_part, mark_spaces=True)}</span>')
        elif op == "insert":
            if after_part:
                after_chunks.append(f'<span class="diff-add">{_visible_text(after_part, mark_spaces=True)}</span>')
        elif op == "replace":
            if before_part:
                before_chunks.append(f'<span class="diff-del">{_visible_text(before_part, mark_spaces=True)}</span>')
            if after_part:
                after_chunks.append(f'<span class="diff-add">{_visible_text(after_part, mark_spaces=True)}</span>')

    before_html = "".join(before_chunks) or "<span class=\"diff-eq\">(empty)</span>"
    after_html = "".join(after_chunks) or "<span class=\"diff-eq\">(empty)</span>"
    return (
        "<div class=\"diff-wrap\">"
        f"<div class=\"diff-line before-line\"><span class=\"diff-label\">Before:</span>{before_html}</div>"
        f"<div class=\"diff-line after-line\"><span class=\"diff-label\">After:</span>{after_html}</div>"
        "<div class=\"diff-note\">Legend: changed regular spaces are marked as ·</div>"
        f"<div class=\"diff-note\">Whitespace delta: {_whitespace_delta_summary(before, after)}</div>"
        "</div>"
    )


def _visible_text(text: str, mark_spaces: bool) -> str:
    out: list[str] = []
    for char in text:
        if mark_spaces and char == " ":
            out.append('<span class="ws ws-space" title="SPACE">·</span>')
        else:
            out.append(escape(char))
    return "".join(out)


def _whitespace_delta_summary(before: str, after: str) -> str:
    before_counter = Counter(before)
    after_counter = Counter(after)
    before_spaces = before_counter.get(" ", 0)
    after_spaces = after_counter.get(" ", 0)
    delta = after_spaces - before_spaces
    if delta == 0:
        return "SPACE: unchanged"
    sign = "+" if delta > 0 else ""
    return f"SPACE: {before_spaces}->{after_spaces} ({sign}{delta})"


def _set_seg_inner_xml(seg: ET.Element, inner_xml: str) -> None:
    replacement_seg = build_seg_from_inner_xml(inner_xml)
    seg.clear()
    seg.text = replacement_seg.text
    for child in list(replacement_seg):
        seg.append(deepcopy(child))


def _verify_cleanup_with_gemini(
    verify_with_gemini: bool,
    gemini_verifier: object | None,
    src_lang: str,
    tgt_lang: str,
    tu_no: int,
    total_tus: int,
    original_src_text: str,
    original_tgt_text: str,
    cleaned_src_text: str,
    cleaned_tgt_text: str,
    cleanup_result: CleanupResult,
    input_price_per_1m: float,
    output_price_per_1m: float,
    gemini_checked: int,
    gemini_rejected: int,
    gemini_input_tokens: int,
    gemini_output_tokens: int,
    gemini_total_tokens: int,
    progress_callback: Callable[[dict[str, object]], None] | None,
    report_items: list[dict[str, object]],
    gemini_audit_events: list[dict[str, object]],
    log: logging.Logger,
) -> dict[str, object] | None:
    if not verify_with_gemini or gemini_verifier is None:
        return None
    if not bool(getattr(gemini_verifier, "supports_cleanup_audit", False)):
        return None
    if not cleanup_result.auto_actions and not cleanup_result.warnings and not cleanup_result.remove_tu:
        return None

    audit_context = {
        "cleanup_actions": cleanup_result.auto_actions,
        "warnings": cleanup_result.warnings,
        "remove_tu": cleanup_result.remove_tu,
        "remove_reason": cleanup_result.remove_reason,
        "original_src": original_src_text,
        "original_tgt": original_tgt_text,
        "cleaned_src": cleaned_src_text,
        "cleaned_tgt": cleaned_tgt_text,
    }
    verify_request = GeminiVerificationRequest(
        src_lang=src_lang,
        tgt_lang=tgt_lang,
        original_src=json.dumps(audit_context, ensure_ascii=False),
        original_tgt=f"TU {tu_no}/{total_tus} cleanup audit",
        src_parts=[cleaned_src_text],
        tgt_parts=[cleaned_tgt_text],
    )
    gemini_checked += 1
    audit_result = _run_gemini_verification(
        gemini_verifier=gemini_verifier,
        verify_request=verify_request,
        prompt_template=GEMINI_CLEANUP_AUDIT_PROMPT,
    )
    if audit_result.verdict == "FAIL":
        gemini_rejected += 1

    gemini_input_tokens += max(0, int(audit_result.prompt_tokens))
    gemini_output_tokens += max(0, int(audit_result.completion_tokens))
    audit_total_tokens = max(
        0,
        int(audit_result.total_tokens),
    )
    if audit_total_tokens == 0:
        audit_total_tokens = max(0, int(audit_result.prompt_tokens)) + max(
            0,
            int(audit_result.completion_tokens),
        )
    gemini_total_tokens += audit_total_tokens
    run_cost = _estimate_cost_usd(
        gemini_input_tokens,
        gemini_output_tokens,
        input_price_per_1m,
        output_price_per_1m,
    )

    log.info(
        (
            "[TU %s/%s] Gemini cleanup audit verdict=%s issues=%s summary=%s "
            "tokens(in=%s out=%s total=%s) run_tokens(in=%s out=%s total=%s) est_cost=$%.6f"
        ),
        tu_no,
        total_tus,
        audit_result.verdict,
        len(audit_result.issues),
        audit_result.summary,
        audit_result.prompt_tokens,
        audit_result.completion_tokens,
        audit_total_tokens,
        gemini_input_tokens,
        gemini_output_tokens,
        gemini_total_tokens,
        run_cost,
    )
    _emit_progress(
        progress_callback,
        {
            "event": "gemini_cleanup_audit",
            "tu_index": tu_no,
            "total_tus": total_tus,
            "verdict": audit_result.verdict,
            "summary": audit_result.summary,
            "issues_count": len(audit_result.issues),
            "gemini_checked": gemini_checked,
            "gemini_rejected": gemini_rejected,
            "gemini_input_tokens": gemini_input_tokens,
            "gemini_output_tokens": gemini_output_tokens,
            "gemini_total_tokens": gemini_total_tokens,
            "gemini_estimated_cost_usd": run_cost,
        },
    )
    report_items.append(
        {
            "category": "cleanup_audit",
            "tu_index": tu_no - 1,
            "verdict": audit_result.verdict,
            "summary": audit_result.summary,
            "issues": [issue.__dict__ for issue in audit_result.issues],
            "prompt_tokens": audit_result.prompt_tokens,
            "completion_tokens": audit_result.completion_tokens,
            "total_tokens": audit_total_tokens,
            "run_gemini_input_tokens": gemini_input_tokens,
            "run_gemini_output_tokens": gemini_output_tokens,
            "run_gemini_total_tokens": gemini_total_tokens,
            "run_estimated_cost_usd": run_cost,
            "cleanup_remove_tu": cleanup_result.remove_tu,
            "cleanup_remove_reason": cleanup_result.remove_reason,
        }
    )
    gemini_audit_events.append(
        {
            "tu_index": tu_no - 1,
            "kind": "cleanup",
            "verdict": audit_result.verdict,
            "summary": audit_result.summary,
            "issues_count": len(audit_result.issues),
            "remove_tu": cleanup_result.remove_tu,
            "remove_reason": cleanup_result.remove_reason,
        }
    )
    return {
        "gemini_checked": gemini_checked,
        "gemini_rejected": gemini_rejected,
        "gemini_input_tokens": gemini_input_tokens,
        "gemini_output_tokens": gemini_output_tokens,
        "gemini_total_tokens": gemini_total_tokens,
        "verdict": audit_result.verdict,
    }


def _build_split_tus(
    tu: ET.Element,
    src_lang: str,
    tgt_lang: str,
    src_parts: list[str],
    tgt_parts: list[str],
    confidence: str,
    gemini_result: GeminiVerificationResult | None,
) -> list[ET.Element]:
    parts_count = len(src_parts)
    split_tus: list[ET.Element] = []

    for idx in range(parts_count):
        new_tu = ET.Element("tu", dict(tu.attrib))
        confidence_prop = ET.Element("prop", {"type": "x-TMXRepair-Confidence"})
        confidence_prop.text = confidence
        new_tu.append(confidence_prop)
        if gemini_result is not None:
            verdict_prop = ET.Element("prop", {"type": "x-TMXRepair-GeminiVerdict"})
            verdict_prop.text = gemini_result.verdict
            new_tu.append(verdict_prop)

        for child in list(tu):
            if _local_name(child.tag) != "tuv":
                if _should_drop_prop_in_split(child):
                    continue
                new_tu.append(deepcopy(child))
                continue

            lang = _get_tuv_lang(child)
            original_seg = child.find("seg")
            if original_seg is None:
                continue

            new_tuv = ET.Element("tuv", dict(child.attrib))
            if lang == src_lang:
                seg_inner = src_parts[idx]
            elif lang == tgt_lang:
                seg_inner = tgt_parts[idx]
            else:
                seg_inner = seg_to_inner_xml(original_seg)

            new_tuv.append(build_seg_from_inner_xml(seg_inner))
            new_tu.append(new_tuv)

        split_tus.append(new_tu)
    return split_tus


def _find_tuv_by_lang(tuv_elements: list[ET.Element], lang: str) -> ET.Element | None:
    return next((tuv for tuv in tuv_elements if _get_tuv_lang(tuv) == lang), None)


def _get_tuv_lang(tuv: ET.Element) -> str:
    return tuv.attrib.get(XML_LANG) or tuv.attrib.get("lang") or ""


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _run_gemini_verification(
    gemini_verifier: object,
    verify_request: GeminiVerificationRequest,
    prompt_template: str | None = None,
) -> GeminiVerificationResult:
    verify_method = getattr(gemini_verifier, "verify_split", None)
    if verify_method is None:
        raise ValueError("gemini_verifier must have verify_split(request) method.")
    if prompt_template is None:
        result = verify_method(verify_request)
    else:
        try:
            result = verify_method(verify_request, prompt_template=prompt_template)
        except TypeError:
            result = verify_method(verify_request)
    if not isinstance(result, GeminiVerificationResult):
        raise ValueError("verify_split(request) must return GeminiVerificationResult.")
    return result


def _write_xlsx_multi_sheet_report(
    path: Path,
    input_path: Path,
    output_path: Path,
    stats: RepairStats,
    split_events: list[dict[str, object]],
    cleanup_events: list[dict[str, object]],
    warning_events: list[dict[str, object]],
    gemini_audit_events: list[dict[str, object]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_fill = PatternFill(start_color="DCEBFF", end_color="DCEBFF", fill_type="solid")
    header_font = Font(bold=True, color="0F172A")
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_rows: list[tuple[str, object]] = [
        ("Input TMX", str(input_path)),
        ("Output TMX", str(output_path)),
        ("Total TU", stats.total_tus),
        ("Split TU", stats.split_tus),
        ("Output TU", stats.created_tus),
        ("Skipped TU", stats.skipped_tus),
        ("High Confidence", stats.high_confidence_splits),
        ("Medium Confidence", stats.medium_confidence_splits),
        ("Gemini Checked", stats.gemini_checked),
        ("Gemini Rejected", stats.gemini_rejected),
        ("Gemini Tokens In", stats.gemini_input_tokens),
        ("Gemini Tokens Out", stats.gemini_output_tokens),
        ("Gemini Tokens Total", stats.gemini_total_tokens),
        ("Gemini Estimated Cost (USD)", f"{stats.gemini_estimated_cost_usd:.6f}"),
        ("AUTO Actions", stats.auto_actions),
        ("AUTO Removed TU", stats.auto_removed_tus),
        ("WARN Issues", stats.warn_issues),
    ]
    summary_ws.append(["Metric", "Value"])
    for key, value in summary_rows:
        summary_ws.append([key, value])
    _style_sheet(summary_ws, header_fill=header_fill, header_font=header_font, cell_alignment=cell_alignment)
    _autosize_columns(summary_ws)

    split_ws = wb.create_sheet("Split Changes")
    split_headers = [
        "TU #",
        "Confidence",
        "Gemini Verdict",
        "Source Before",
        "Target Before",
        "Source After Parts",
        "Target After Parts",
    ]
    split_ws.append(split_headers)
    for event in split_events:
        split_ws.append(
            [
                int(event.get("tu_index", 0)) + 1,
                str(event.get("confidence", "")),
                str(event.get("gemini_verdict", "") or ""),
                str(event.get("original_src", "")),
                str(event.get("original_tgt", "")),
                _join_parts_for_sheet(event.get("src_parts", [])),
                _join_parts_for_sheet(event.get("tgt_parts", [])),
            ]
        )
    _style_sheet(split_ws, header_fill=header_fill, header_font=header_font, cell_alignment=cell_alignment)
    _autosize_columns(split_ws)

    cleanup_ws = wb.create_sheet("Auto Cleanup")
    cleanup_headers = [
        "TU #",
        "Rule",
        "Message",
        "Source Before",
        "Source After",
        "Target Before",
        "Target After",
        "Remove Reason",
    ]
    cleanup_ws.append(cleanup_headers)
    for event in cleanup_events:
        cleanup_ws.append(
            [
                int(event.get("tu_index", 0)) + 1,
                str(event.get("rule", "")),
                str(event.get("message", "")),
                str(event.get("before_src", "")),
                str(event.get("after_src", "")),
                str(event.get("before_tgt", "")),
                str(event.get("after_tgt", "")),
                str(event.get("remove_reason", "") or ""),
            ]
        )
    _style_sheet(cleanup_ws, header_fill=header_fill, header_font=header_font, cell_alignment=cell_alignment)
    _autosize_columns(cleanup_ws)

    warnings_ws = wb.create_sheet("Warnings")
    warning_headers = [
        "TU #",
        "Rule",
        "Severity",
        "Message",
        "Source Snapshot",
        "Target Snapshot",
    ]
    warnings_ws.append(warning_headers)
    for event in warning_events:
        warnings_ws.append(
            [
                int(event.get("tu_index", 0)) + 1,
                str(event.get("rule", "")),
                str(event.get("severity", "")),
                str(event.get("message", "")),
                str(event.get("src_text", "")),
                str(event.get("tgt_text", "")),
            ]
        )
    _style_sheet(warnings_ws, header_fill=header_fill, header_font=header_font, cell_alignment=cell_alignment)
    _autosize_columns(warnings_ws)

    gemini_ws = wb.create_sheet("Gemini Checks")
    gemini_headers = [
        "TU #",
        "Kind",
        "Verdict",
        "Issues Count",
        "Summary",
        "Remove TU",
        "Remove Reason",
    ]
    gemini_ws.append(gemini_headers)
    for event in gemini_audit_events:
        gemini_ws.append(
            [
                int(event.get("tu_index", 0)) + 1,
                str(event.get("kind", "")),
                str(event.get("verdict", "")),
                int(event.get("issues_count", 0) or 0),
                str(event.get("summary", "")),
                str(event.get("remove_tu", "")),
                str(event.get("remove_reason", "") or ""),
            ]
        )
    _style_sheet(gemini_ws, header_fill=header_fill, header_font=header_font, cell_alignment=cell_alignment)
    _autosize_columns(gemini_ws)

    wb.save(path)


def _style_sheet(
    worksheet: object,
    *,
    header_fill: object,
    header_font: object,
    cell_alignment: object,
) -> None:
    rows = list(worksheet.iter_rows())  # type: ignore[attr-defined]
    if not rows:
        return
    for cell in rows[0]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows[1:]:
        for cell in row:
            cell.alignment = cell_alignment


def _autosize_columns(worksheet: object, max_width: int = 80) -> None:
    for column_cells in worksheet.columns:  # type: ignore[attr-defined]
        length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            length = max(length, len(str(value)))
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(12, length + 2), max_width)  # type: ignore[attr-defined]


def _join_parts_for_sheet(parts: object) -> str:
    if isinstance(parts, list):
        return "\n".join(f"{idx}. {str(part)}" for idx, part in enumerate(parts, start=1))
    return str(parts)


def _write_html_diff_report(
    path: Path,
    input_path: Path,
    output_path: Path,
    stats: RepairStats,
    split_events: list[dict[str, object]],
    cleanup_events: list[dict[str, object]],
    warning_events: list[dict[str, object]],
    gemini_audit_events: list[dict[str, object]],
) -> None:
    split_blocks: list[str] = []
    for event in split_events:
        src_parts_html = "".join(f"<li>{escape(part)}</li>" for part in event["src_parts"])
        tgt_parts_html = "".join(f"<li>{escape(part)}</li>" for part in event["tgt_parts"])
        confidence = escape(str(event["confidence"]))
        gemini_verdict = event.get("gemini_verdict")
        gemini_badge = ""
        if gemini_verdict is not None:
            gemini_badge = f'<span class="badge">Gemini: {escape(str(gemini_verdict))}</span>'

        split_blocks.append(
            f"""
            <section class="card">
              <h2>TU #{int(event["tu_index"]) + 1}</h2>
              <p><span class="badge">Confidence: {confidence}</span>{gemini_badge}</p>
              <div class="grid">
                <div class="pane">
                  <h3>Source Before</h3>
                  <pre>{escape(str(event["original_src"]))}</pre>
                  <h3>Source After</h3>
                  <ol>{src_parts_html}</ol>
                </div>
                <div class="pane">
                  <h3>Target Before</h3>
                  <pre>{escape(str(event["original_tgt"]))}</pre>
                  <h3>Target After</h3>
                  <ol>{tgt_parts_html}</ol>
                </div>
              </div>
            </section>
            """
        )

    if not split_blocks:
        split_blocks.append(
            '<section class="card"><h2>No Split Changes</h2>'
            "<p>No TU entries were split in this run.</p></section>"
        )

    cleanup_blocks: list[str] = []
    for event in cleanup_events:
        before_src = str(event.get("before_src", ""))
        after_src = str(event.get("after_src", ""))
        before_tgt = str(event.get("before_tgt", ""))
        after_tgt = str(event.get("after_tgt", ""))
        src_diff_html = _render_inline_diff(before_src, after_src)
        tgt_diff_html = _render_inline_diff(before_tgt, after_tgt)
        cleanup_blocks.append(
            f"""
            <section class="card">
              <h2>TU #{int(event.get("tu_index", 0)) + 1}</h2>
              <p><span class="badge auto">AUTO: {escape(str(event.get("rule", "")))}</span></p>
              <p>{escape(str(event.get("message", "")))}</p>
              <div class="grid">
                <div class="pane">
                  <h3>Source Diff</h3>
                  {src_diff_html}
                </div>
                <div class="pane">
                  <h3>Target Diff</h3>
                  {tgt_diff_html}
                </div>
              </div>
            </section>
            """
        )
    if not cleanup_blocks:
        cleanup_blocks.append(
            '<section class="card"><h2>No Auto Cleanup Actions</h2>'
            "<p>No AUTO cleanup actions were applied in this run.</p></section>"
        )

    warning_blocks: list[str] = []
    for event in warning_events:
        warning_blocks.append(
            f"""
            <section class="card">
              <h2>TU #{int(event.get("tu_index", 0)) + 1}</h2>
              <p><span class="badge warn">{escape(str(event.get("rule", "")))}</span></p>
              <p>{escape(str(event.get("message", "")))}</p>
              <div class="grid">
                <div class="pane">
                  <h3>Source Snapshot</h3>
                  <pre>{escape(str(event.get("src_text", "")))}</pre>
                </div>
                <div class="pane">
                  <h3>Target Snapshot</h3>
                  <pre>{escape(str(event.get("tgt_text", "")))}</pre>
                </div>
              </div>
            </section>
            """
        )
    if not warning_blocks:
        warning_blocks.append(
            '<section class="card"><h2>No Warnings</h2>'
            "<p>No WARN diagnostics were produced in this run.</p></section>"
        )

    gemini_blocks: list[str] = []
    for event in gemini_audit_events:
        kind = escape(str(event.get("kind", "unknown")))
        verdict = escape(str(event.get("verdict", "n/a")))
        summary = escape(str(event.get("summary", "")))
        issues_count = int(event.get("issues_count", 0) or 0)
        gemini_blocks.append(
            f"""
            <section class="card">
              <h2>TU #{int(event.get("tu_index", 0)) + 1}</h2>
              <p><span class="badge">Gemini {kind}</span><span class="badge">{verdict}</span></p>
              <p>Issues: {issues_count}</p>
              <p>{summary}</p>
            </section>
            """
        )
    if not gemini_blocks:
        gemini_blocks.append(
            '<section class="card"><h2>No Gemini Checks</h2>'
            "<p>Gemini verification was not used for this run.</p></section>"
        )

    html_text = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>TMX Repair Diff Report</title>
  <style>
    body {{ font-family: Segoe UI, Tahoma, Arial, sans-serif; background: #f5f7fa; color: #1f2937; margin: 0; }}
    .wrap {{ max-width: 1200px; margin: 0 auto; padding: 20px; }}
    .hero {{ background: linear-gradient(135deg, #0f766e, #2563eb); color: #fff; border-radius: 12px; padding: 18px 20px; }}
    .meta {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 8px 16px; margin-top: 10px; }}
    .card {{ background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; padding: 16px; margin-top: 14px; box-shadow: 0 1px 2px rgba(0,0,0,0.04); }}
    .grid {{ display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 14px; }}
    .pane {{ background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 10px; padding: 10px; }}
    .tabs {{ display: flex; gap: 8px; flex-wrap: wrap; margin-top: 14px; }}
    .tab-button {{ border: 1px solid #cbd5e1; border-radius: 10px; background: #fff; padding: 8px 12px; cursor: pointer; font-weight: 600; }}
    .tab-button.active {{ background: #0f766e; color: #fff; border-color: #0f766e; }}
    .tab-panel {{ display: none; }}
    .tab-panel.active {{ display: block; }}
    .badge {{ display: inline-block; background: #e0f2fe; color: #075985; border-radius: 999px; padding: 4px 10px; margin-right: 8px; font-size: 12px; font-weight: 600; }}
    .badge.auto {{ background: #dcfce7; color: #166534; }}
    .badge.warn {{ background: #fef3c7; color: #92400e; }}
    .diff-wrap {{ display: grid; gap: 8px; }}
    .diff-line {{ background: #fff; border: 1px dashed #cbd5e1; border-radius: 8px; padding: 8px; overflow-x: auto; white-space: pre-wrap; word-break: break-word; font-family: Consolas, 'Courier New', monospace; font-size: 13px; }}
    .diff-label {{ font-weight: 700; margin-right: 6px; color: #334155; }}
    .diff-del {{ background: #fee2e2; color: #991b1b; border-radius: 4px; padding: 0 1px; }}
    .diff-add {{ background: #86efac; color: #14532d; border: 1px solid #15803d; border-radius: 4px; padding: 0 1px; font-weight: 700; }}
    .diff-eq {{ color: #1f2937; }}
    .after-line .diff-eq {{ color: #475569; }}
    .diff-note {{ font-size: 12px; color: #64748b; margin-top: 4px; }}
    .ws {{ display: inline-block; min-width: 0.8em; text-align: center; border-radius: 3px; border: 1px solid #93c5fd; margin: 0 0.5px; font-size: 11px; line-height: 1.05; background: #dbeafe; color: #1d4ed8; }}
    .ws-space {{ background: #dbeafe; }}
    pre {{ white-space: pre-wrap; word-break: break-word; background: #fff; border: 1px dashed #cbd5e1; border-radius: 8px; padding: 8px; }}
    ol {{ margin: 0; padding-left: 20px; }}
    li {{ margin: 4px 0; }}
    @media (max-width: 900px) {{ .grid {{ grid-template-columns: 1fr; }} }}
  </style>
</head>
<body>
  <div class="wrap">
    <header class="hero">
      <h1>TMX Repair Diff Report</h1>
      <div class="meta">
        <div>Input: {escape(str(input_path))}</div>
        <div>Output: {escape(str(output_path))}</div>
        <div>Total TU: {stats.total_tus}</div>
        <div>Split TU: {stats.split_tus}</div>
        <div>Output TU: {stats.created_tus}</div>
        <div>Skipped TU: {stats.skipped_tus}</div>
        <div>High Confidence: {stats.high_confidence_splits}</div>
        <div>Medium Confidence: {stats.medium_confidence_splits}</div>
        <div>Gemini Checked: {stats.gemini_checked}</div>
        <div>Gemini Rejected: {stats.gemini_rejected}</div>
        <div>Gemini Tokens In: {stats.gemini_input_tokens}</div>
        <div>Gemini Tokens Out: {stats.gemini_output_tokens}</div>
        <div>Gemini Tokens Total: {stats.gemini_total_tokens}</div>
        <div>Gemini Est Cost (USD): {stats.gemini_estimated_cost_usd:.6f}</div>
        <div>AUTO Actions: {stats.auto_actions}</div>
        <div>AUTO Removed TU: {stats.auto_removed_tus}</div>
        <div>WARN Issues: {stats.warn_issues}</div>
      </div>
    </header>
    <div class="tabs">
      <button class="tab-button active" data-tab="split">Split Changes</button>
      <button class="tab-button" data-tab="cleanup">Auto Cleanup</button>
      <button class="tab-button" data-tab="warnings">Warnings</button>
      <button class="tab-button" data-tab="gemini">Gemini Checks</button>
    </div>
    <section class="tab-panel active" id="tab-split">{"".join(split_blocks)}</section>
    <section class="tab-panel" id="tab-cleanup">{"".join(cleanup_blocks)}</section>
    <section class="tab-panel" id="tab-warnings">{"".join(warning_blocks)}</section>
    <section class="tab-panel" id="tab-gemini">{"".join(gemini_blocks)}</section>
  </div>
  <script>
    (function() {{
      var buttons = document.querySelectorAll(".tab-button");
      var panels = document.querySelectorAll(".tab-panel");
      buttons.forEach(function(btn) {{
        btn.addEventListener("click", function() {{
          buttons.forEach(function(other) {{ other.classList.remove("active"); }});
          panels.forEach(function(panel) {{ panel.classList.remove("active"); }});
          btn.classList.add("active");
          var target = document.getElementById("tab-" + btn.getAttribute("data-tab"));
          if (target) {{
            target.classList.add("active");
          }}
        }});
      }});
    }})();
  </script>
</body>
</html>
"""
    path.write_text(html_text, encoding="utf-8")
