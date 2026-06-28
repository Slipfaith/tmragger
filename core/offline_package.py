"""Offline .tmrepair package export/import helpers."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from difflib import SequenceMatcher
import hashlib
import io
import json
from pathlib import Path
import re
import tempfile
from typing import Any, Callable
import zipfile

from core.plan import Proposal, RepairPlan

FORMAT_VERSION = 1
MANIFEST_NAME = "manifest.json"
STATE_NAME = "state.json"
SOURCE_NAME = "source.tmx"
REPORT_XLSX_NAME = "report.xlsx"
DECISIONS_NAME = "decisions.json"

STATUS_PENDING = "pending"
STATUS_ACCEPTED = "accepted"
STATUS_REJECTED = "rejected"
STATUS_SKIPPED = "skipped"

_STATUS_ORDER = {
    STATUS_ACCEPTED: 0,
    STATUS_REJECTED: 1,
    STATUS_SKIPPED: 2,
    STATUS_PENDING: 3,
}

_DECISION_TO_STATUS = {
    "accept": STATUS_ACCEPTED,
    "accepted": STATUS_ACCEPTED,
    "approve": STATUS_ACCEPTED,
    "approved": STATUS_ACCEPTED,
    "reject": STATUS_REJECTED,
    "rejected": STATUS_REJECTED,
    "decline": STATUS_REJECTED,
    "skip": STATUS_SKIPPED,
    "skipped": STATUS_SKIPPED,
    "pending": STATUS_PENDING,
}

_SERVICE_MARKUP_RULES = {
    "remove_inline_tags",
    "remove_game_markup",
    "remove_percent_wrapped_tokens",
    "context_remove_game_markup",
    "context_remove_percent_wrapped_tokens",
}

_TYPE_LABELS = {
    "split": "Split",
    "dedup_tu": "Удаление дублей",
    "normalize_spaces": "Очистка пробелов",
    "remove_line_breaks": "Переносы строк",
    "service_markup": "Удаление служебной разметки",
    "remove_garbage_segment": "Удаление мусорных TU",
}
@dataclass(slots=True)
class PackageImportResult:
    package_path: Path
    manifest: dict[str, Any]
    state: dict[str, Any]
    plan: RepairPlan
    accepted_count: int
    rejected_count: int
    skipped_count: int
    unrecognized_count: int
    decisions_source: str
    source_tmx_path: Path
    hash_mismatch_warning: str | None


def export_tmrepair_package(
    package_path: Path,
    input_tmx_path: Path,
    plan: RepairPlan,
    settings: dict[str, Any] | None = None,
    progress_callback: Callable[[int, int], None] | None = None,
) -> None:
    """Export a self-contained .tmrepair zip package.

    ``progress_callback(done_rows, total_rows)`` is invoked while the (slow)
    XLSX report is being built, so callers can show a live counter.
    """
    source_bytes = input_tmx_path.read_bytes()
    package_id = _sha256_bytes(source_bytes)[:16]
    now_iso = _utc_now_iso()
    state = _state_from_plan(plan=plan, package_id=package_id, timestamp_iso=now_iso)
    manifest = {
        "format_version": FORMAT_VERSION,
        "package_id": package_id,
        "source_file_name": input_tmx_path.name,
        "source_sha256": _sha256_bytes(source_bytes),
        "created_at": now_iso,
        "updated_at": now_iso,
        "settings": settings or {},
    }

    xlsx_bytes = _build_xlsx_report(
        state=state, manifest=manifest, progress_callback=progress_callback
    )

    # The XLSX is the single editing surface: it carries every issue with no
    # row cap and the importer reads decisions back from it. A standalone HTML
    # report was dropped because large plans produced multi-MB single-page
    # documents that browsers struggled to open.
    _write_zip_atomically(
        package_path=package_path,
        replacements={
            MANIFEST_NAME: _json_bytes(manifest),
            STATE_NAME: _json_bytes(state),
            SOURCE_NAME: source_bytes,
            REPORT_XLSX_NAME: xlsx_bytes,
        },
    )


def import_tmrepair_package(
    package_path: Path,
    external_decisions_path: Path | None = None,
) -> PackageImportResult:
    """Import decisions from .tmrepair package and return an updated plan."""
    with zipfile.ZipFile(package_path, "r") as archive:
        manifest = json.loads(archive.read(MANIFEST_NAME).decode("utf-8"))
        state = json.loads(archive.read(STATE_NAME).decode("utf-8"))
        source_bytes = archive.read(SOURCE_NAME)
        decisions_payload, decisions_source = _load_decisions_payload(
            archive=archive,
            external_decisions_path=external_decisions_path,
        )

    unrecognized_count = _apply_decisions_to_state(state=state, decisions_payload=decisions_payload)
    state["updated_at"] = _utc_now_iso()
    manifest["updated_at"] = state["updated_at"]

    _write_zip_atomically(
        package_path=package_path,
        replacements={
            MANIFEST_NAME: _json_bytes(manifest),
            STATE_NAME: _json_bytes(state),
        },
    )

    source_temp = _write_source_temp_file(
        package_path=package_path,
        source_name=str(manifest.get("source_file_name", "source.tmx")),
        source_bytes=source_bytes,
    )

    plan = _plan_from_state(state=state, input_path=str(manifest.get("source_file_name", "")))
    accepted_count, rejected_count, skipped_count = _status_counts(state)
    hash_warning = _build_hash_warning(package_path=package_path, manifest=manifest)

    return PackageImportResult(
        package_path=package_path,
        manifest=manifest,
        state=state,
        plan=plan,
        accepted_count=accepted_count,
        rejected_count=rejected_count,
        skipped_count=skipped_count,
        unrecognized_count=unrecognized_count,
        decisions_source=decisions_source,
        source_tmx_path=source_temp,
        hash_mismatch_warning=hash_warning,
    )


def _write_source_temp_file(package_path: Path, source_name: str, source_bytes: bytes) -> Path:
    source_suffix = Path(source_name).suffix or ".tmx"
    fd, temp_name = tempfile.mkstemp(
        prefix=f"{package_path.stem}_source_",
        suffix=source_suffix,
        dir=str(package_path.parent),
    )
    temp_path = Path(temp_name)
    try:
        with io.FileIO(fd, "wb", closefd=True) as stream:
            stream.write(source_bytes)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
    return temp_path


def _load_decisions_payload(
    archive: zipfile.ZipFile,
    external_decisions_path: Path | None,
) -> tuple[list[dict[str, str]], str]:
    names = set(archive.namelist())
    if DECISIONS_NAME in names:
        payload = archive.read(DECISIONS_NAME).decode("utf-8")
        return _parse_decisions_json(payload), DECISIONS_NAME

    if REPORT_XLSX_NAME in names:
        return _parse_decisions_xlsx(archive.read(REPORT_XLSX_NAME)), REPORT_XLSX_NAME

    if external_decisions_path is not None and external_decisions_path.exists():
        suffix = external_decisions_path.suffix.lower()
        if suffix == ".json":
            return _parse_decisions_json(external_decisions_path.read_text(encoding="utf-8")), str(
                external_decisions_path
            )
        if suffix in {".xlsx", ".xlsm"}:
            return _parse_decisions_xlsx(external_decisions_path.read_bytes()), str(external_decisions_path)
    return [], "none"


def _parse_decisions_json(payload: str) -> list[dict[str, str]]:
    try:
        raw = json.loads(payload)
    except json.JSONDecodeError:
        return []
    decisions_raw: list[Any]
    if isinstance(raw, dict):
        decisions_raw = list(raw.get("decisions", []))
    elif isinstance(raw, list):
        decisions_raw = raw
    else:
        return []

    parsed: list[dict[str, str]] = []
    for item in decisions_raw:
        if not isinstance(item, dict):
            continue
        issue_id = str(item.get("id", "")).strip()
        if not issue_id:
            continue
        decision = str(item.get("decision", "")).strip().lower()
        comment = str(item.get("comment", "")).strip()
        parsed.append({"id": issue_id, "decision": decision, "comment": comment})
    return parsed


def _parse_decisions_xlsx(payload: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(payload), data_only=True)
    try:
        worksheet = workbook["Review"] if "Review" in workbook.sheetnames else workbook.active
        header_row = None
        header_map: dict[str, int] = {}
        for row_index in range(1, min(30, worksheet.max_row) + 1):
            values = [str(worksheet.cell(row=row_index, column=col).value or "").strip() for col in range(1, 20)]
            if "ID проблемы" in values and "Решение" in values:
                header_row = row_index
                for col_index, value in enumerate(values, start=1):
                    if value:
                        header_map[value] = col_index
                break
        if header_row is None:
            return []

        id_col = header_map.get("ID проблемы")
        decision_col = header_map.get("Решение")
        comment_col = header_map.get("Комментарий")
        if id_col is None or decision_col is None:
            return []

        decisions: list[dict[str, str]] = []
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            issue_id = str(worksheet.cell(row=row_index, column=id_col).value or "").strip()
            if not issue_id:
                continue
            decision = str(worksheet.cell(row=row_index, column=decision_col).value or "").strip().lower()
            comment = ""
            if comment_col is not None:
                comment = str(worksheet.cell(row=row_index, column=comment_col).value or "").strip()
            decisions.append({"id": issue_id, "decision": decision, "comment": comment})
        return decisions
    finally:
        workbook.close()


def _apply_decisions_to_state(state: dict[str, Any], decisions_payload: list[dict[str, str]]) -> int:
    issues = state.get("issues", [])
    if not isinstance(issues, list):
        return 0
    by_id: dict[str, dict[str, Any]] = {}
    for issue in issues:
        if not isinstance(issue, dict):
            continue
        issue_id = str(issue.get("id", "")).strip()
        if issue_id:
            by_id[issue_id] = issue

    unrecognized = 0
    for decision in decisions_payload:
        issue_id = str(decision.get("id", "")).strip()
        if not issue_id:
            continue
        issue = by_id.get(issue_id)
        if issue is None:
            unrecognized += 1
            continue
        normalized = _normalize_status(decision.get("decision", ""))
        if normalized is None:
            unrecognized += 1
            continue
        issue["status"] = normalized
        comment = str(decision.get("comment", "")).strip()
        if comment:
            issue["comment"] = comment
    return unrecognized


def _normalize_status(value: str | Any) -> str | None:
    normalized = str(value or "").strip().lower()
    if not normalized:
        return None
    return _DECISION_TO_STATUS.get(normalized)


def _status_counts(state: dict[str, Any]) -> tuple[int, int, int]:
    accepted = 0
    rejected = 0
    skipped = 0
    for issue in state.get("issues", []):
        if not isinstance(issue, dict):
            continue
        status = str(issue.get("status", "")).strip().lower()
        if status == STATUS_ACCEPTED:
            accepted += 1
        elif status == STATUS_REJECTED:
            rejected += 1
        elif status == STATUS_SKIPPED:
            skipped += 1
    return accepted, rejected, skipped


def _build_hash_warning(package_path: Path, manifest: dict[str, Any]) -> str | None:
    source_name = str(manifest.get("source_file_name", "")).strip()
    source_hash = str(manifest.get("source_sha256", "")).strip().lower()
    if not source_name or not source_hash:
        return None
    disk_path = package_path.parent / source_name
    if not disk_path.exists():
        return None
    disk_hash = _sha256_file(disk_path)
    if disk_hash.lower() == source_hash:
        return None
    return (
        f"TMX hash mismatch for {disk_path.name}: "
        f"package={source_hash}, disk={disk_hash}. "
        "Package source will be used for apply."
    )


def _state_from_plan(plan: RepairPlan, package_id: str, timestamp_iso: str) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    for proposal in plan.proposals:
        issues.append(_issue_from_proposal(proposal))
    issues.sort(key=lambda item: (_STATUS_ORDER.get(str(item.get("status", "")), 99), int(item.get("tu_index", 0))))
    return {
        "package_id": package_id,
        "created_at": timestamp_iso,
        "updated_at": timestamp_iso,
        "total_tus": int(plan.total_tus),
        "issues": issues,
    }


def _issue_from_proposal(proposal: Proposal) -> dict[str, Any]:
    check_type = _proposal_check_type(proposal)
    return {
        "id": proposal.proposal_id,
        "kind": proposal.kind,
        "check_type": check_type,
        "check_type_label": _type_label(check_type),
        "tu_index": int(proposal.tu_index),
        "status": STATUS_PENDING,
        "comment": "",
        "confidence": proposal.confidence,
        "gemini_verdict": proposal.gemini_verdict,
        "rule": proposal.rule,
        "message": proposal.message,
        "before_src": proposal.before_src,
        "after_src": proposal.after_src,
        "before_tgt": proposal.before_tgt,
        "after_tgt": proposal.after_tgt,
        "original_src": proposal.original_src,
        "original_tgt": proposal.original_tgt,
        "src_parts": list(proposal.src_parts),
        "tgt_parts": list(proposal.tgt_parts),
    }


def _proposal_check_type(proposal: Proposal) -> str:
    if proposal.kind == "split":
        return "split"
    rule = proposal.rule or "cleanup"
    if rule in _SERVICE_MARKUP_RULES:
        return "service_markup"
    return rule


def _type_label(check_type: str) -> str:
    return _TYPE_LABELS.get(check_type, check_type)


def _plan_from_state(state: dict[str, Any], input_path: str) -> RepairPlan:
    proposals: list[Proposal] = []
    for issue in state.get("issues", []):
        if not isinstance(issue, dict):
            continue
        status = str(issue.get("status", STATUS_PENDING)).strip().lower()
        accepted = status == STATUS_ACCEPTED
        kind = str(issue.get("kind", "cleanup")).strip() or "cleanup"
        if kind not in {"split", "cleanup"}:
            kind = "cleanup"
        proposals.append(
            Proposal(
                proposal_id=str(issue.get("id", "")),
                kind=kind,
                tu_index=int(issue.get("tu_index", 0) or 0),
                accepted=accepted,
                confidence=str(issue.get("confidence", "")),
                gemini_verdict=str(issue.get("gemini_verdict", "")),
                src_parts=list(issue.get("src_parts", []) or []),
                tgt_parts=list(issue.get("tgt_parts", []) or []),
                rule=str(issue.get("rule", "")),
                message=str(issue.get("message", "")),
                before_src=str(issue.get("before_src", "")),
                after_src=str(issue.get("after_src", "")),
                before_tgt=str(issue.get("before_tgt", "")),
                after_tgt=str(issue.get("after_tgt", "")),
                original_src=str(issue.get("original_src", "")),
                original_tgt=str(issue.get("original_tgt", "")),
            )
        )
    total_tus = int(state.get("total_tus", 0) or 0)
    if total_tus <= 0 and proposals:
        total_tus = max(p.tu_index for p in proposals) + 1
    return RepairPlan(input_path=input_path, total_tus=total_tus, proposals=proposals)


# OOXML forbids the C0 control characters (except tab/newline/CR) even inside a
# cell's string content, and openpyxl's rich-text writer does not strip them.
# Segment text carrying such bytes otherwise yields a workbook Excel reports as
# corrupt (or makes the lxml backend raise). Clean it before writing.
_ILLEGAL_XLSX_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _xlsx_safe_text(text: str) -> str:
    return _ILLEGAL_XLSX_CHARS_RE.sub("", text)


def _build_xlsx_report(
    state: dict[str, Any],
    manifest: dict[str, Any],
    progress_callback: Callable[[int, int], None] | None = None,
) -> bytes:
    # XlsxWriter is write-only but produces Excel-clean files quickly even for
    # 100k-row reports; the importer still reads decisions back via openpyxl.
    import xlsxwriter

    output = io.BytesIO()
    # in_memory avoids temp files (and a Windows atexit cleanup race); the shared
    # string table it keeps also dedupes the report's many repeated segments.
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    sheet = workbook.add_worksheet("Review")

    header_fmt = workbook.add_format(
        {"bold": True, "font_color": "#0F172A", "bg_color": "#DCEBFF",
         "valign": "top", "text_wrap": True}
    )
    cell_fmt = workbook.add_format({"valign": "top", "text_wrap": True})
    editable_fmt = workbook.add_format(
        {"valign": "top", "text_wrap": True, "locked": False}
    )
    # Solid #RRGGBB colors — no alpha, so whole-deletion rows stay visible.
    normal_run = workbook.add_format()
    deleted_run = workbook.add_format({"font_color": "#B91C1C", "font_strikeout": True})
    added_run = workbook.add_format({"font_color": "#15803D"})
    # Single-fragment cells can't use write_rich_string, so combine run style
    # with the cell style (wrap/valign) for those.
    del_cell_fmt = workbook.add_format(
        {"valign": "top", "text_wrap": True, "font_color": "#B91C1C", "font_strikeout": True}
    )
    add_cell_fmt = workbook.add_format(
        {"valign": "top", "text_wrap": True, "font_color": "#15803D"}
    )
    run_fmt = {"eq": normal_run, "del": deleted_run, "add": added_run}
    single_fmt = {"eq": cell_fmt, "del": del_cell_fmt, "add": add_cell_fmt}

    def _diff_runs(
        before_text: str, after_text: str, *, whitespace_focused: bool
    ) -> list[tuple[str, str]]:
        # Whole-deletion (dedup/garbage, ~the bulk of large plans) and pure
        # additions don't need a char-level diff — skip the SequenceMatcher.
        if not after_text:
            text = before_text.replace(" ", "·") if whitespace_focused else before_text
            return [("del", text)] if text else []
        if not before_text:
            text = after_text.replace(" ", "·") if whitespace_focused else after_text
            return [("add", text)] if text else []
        before_segments, after_segments = _diff_segments(
            before_text, after_text, by_char=True
        )
        runs: list[tuple[str, str]] = []

        def _add(kind: str, text: str) -> None:
            if text:
                runs.append((kind, text))

        bi = ai = 0
        bl, al = len(before_segments), len(after_segments)
        while bi < bl or ai < al:
            if bi < bl:
                bk, bt = before_segments[bi]
                if bk == "eq":
                    _add("eq", bt)
                    bi += 1
                    ai += 1
                    continue
                if bk == "del":
                    _add("del", bt.replace(" ", "·") if whitespace_focused else bt)
                    bi += 1
                    continue
                bi += 1
            if ai < al:
                ak, at = after_segments[ai]
                if ak == "add":
                    _add("add", at.replace(" ", "·") if whitespace_focused else at)
                    ai += 1
                    continue
                if ak == "eq":
                    _add("eq", at)
                    ai += 1
                    continue
                ai += 1
        return runs

    def _write_review_cell(
        row: int, col: int, before_text: str, after_text: str, *, whitespace_focused: bool
    ) -> None:
        # Empty segments (dedup/garbage) get a marker so the cell isn't blank.
        if not before_text.strip() and not after_text.strip():
            sheet.write_string(row, col, "(пусто)", cell_fmt)
            return
        runs = _diff_runs(before_text, after_text, whitespace_focused=whitespace_focused)
        if len(runs) <= 1:
            kind, text = runs[0] if runs else ("eq", "")
            sheet.write_string(row, col, text, single_fmt[kind])
            return
        parts: list = []
        for kind, text in runs:
            parts.append(run_fmt[kind])
            parts.append(text)
        sheet.write_rich_string(row, col, *parts, cell_fmt)

    widths = [34, 28, 46, 46, 42, 14, 36]
    for col_index, width in enumerate(widths):
        sheet.set_column(col_index, col_index, width)

    header_row = 3  # 0-based; row 4 in Excel
    start_data_row = header_row + 1
    sheet.freeze_panes(start_data_row, 0)

    sheet.write_string(0, 0, "tmrepair_package_id", cell_fmt)
    sheet.write_string(0, 1, str(manifest.get("package_id", "")), cell_fmt)
    sheet.write_string(1, 0, "tmx_sha256", cell_fmt)
    sheet.write_string(1, 1, str(manifest.get("source_sha256", "")), cell_fmt)
    sheet.write_string(2, 0, "source_file_name", cell_fmt)
    sheet.write_string(2, 1, str(manifest.get("source_file_name", "")), cell_fmt)

    headers = [
        "ID проблемы",
        "Тип проверки",
        "TU source",
        "TU target",
        "Описание",
        "Решение",
        "Комментарий",
    ]
    for col_index, title in enumerate(headers):
        sheet.write_string(header_row, col_index, title, header_fmt)

    issues = state.get("issues", [])
    total_issues = len(issues) if isinstance(issues, list) else 0
    last_row = header_row
    for row_offset, issue in enumerate(issues):
        row = start_data_row + row_offset
        if progress_callback is not None and row_offset % 200 == 0:
            progress_callback(row_offset, total_issues)
        if not isinstance(issue, dict):
            continue
        before_src, after_src, before_tgt, after_tgt = _issue_before_after_text(issue)
        before_src = _xlsx_safe_text(before_src)
        after_src = _xlsx_safe_text(after_src)
        before_tgt = _xlsx_safe_text(before_tgt)
        after_tgt = _xlsx_safe_text(after_tgt)
        check_type = str(issue.get("check_type", "")).strip()
        whitespace_focused = check_type == "normalize_spaces"

        sheet.write_string(row, 0, str(issue.get("id", "")), cell_fmt)
        sheet.write_string(
            row, 1, str(issue.get("check_type_label", issue.get("check_type", ""))), cell_fmt
        )
        _write_review_cell(row, 2, before_src, after_src, whitespace_focused=whitespace_focused)
        _write_review_cell(row, 3, before_tgt, after_tgt, whitespace_focused=whitespace_focused)
        sheet.write_string(row, 4, str(issue.get("message", "")), cell_fmt)
        sheet.write_string(row, 5, "pending", editable_fmt)
        sheet.write_string(row, 6, str(issue.get("comment", "")), editable_fmt)
        last_row = row

    if progress_callback is not None:
        progress_callback(total_issues, total_issues)

    if last_row >= start_data_row:
        sheet.data_validation(
            start_data_row, 5, last_row, 5,
            {"validate": "list", "source": ["accept", "reject", "skip", "pending"]},
        )

    sheet.autofilter(header_row, 0, max(header_row, last_row), len(headers) - 1)
    # Lock the sheet (Decision/Comment cells are unlocked via editable_fmt) but
    # keep the header filter and sorting usable.
    sheet.protect("", {"autofilter": True, "sort": True})

    workbook.close()
    return output.getvalue()


def _issue_before_after_text(issue: dict[str, Any]) -> tuple[str, str, str, str]:
    kind = str(issue.get("kind", "")).strip()
    if kind == "split":
        before_src = str(issue.get("original_src", ""))
        before_tgt = str(issue.get("original_tgt", ""))
        after_src = "\n".join(str(part) for part in list(issue.get("src_parts", []) or []))
        after_tgt = "\n".join(str(part) for part in list(issue.get("tgt_parts", []) or []))
        return before_src, after_src, before_tgt, after_tgt
    before_src = str(issue.get("before_src", "") or issue.get("original_src", ""))
    after_src = str(issue.get("after_src", ""))
    before_tgt = str(issue.get("before_tgt", "") or issue.get("original_tgt", ""))
    after_tgt = str(issue.get("after_tgt", ""))
    return before_src, after_src, before_tgt, after_tgt


def _tokenize_diff_text(text: str) -> list[str]:
    if not text:
        return []
    return re.findall(r"\w+|\s+|[^\w\s]", text, flags=re.UNICODE)


def _diff_segments(
    before_text: str, after_text: str, *, by_char: bool = False
) -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    if by_char:
        before_units = list(before_text)
        after_units = list(after_text)
    else:
        before_units = _tokenize_diff_text(before_text)
        after_units = _tokenize_diff_text(after_text)
    matcher = SequenceMatcher(a=before_units, b=after_units, autojunk=False)
    before_segments: list[tuple[str, str]] = []
    after_segments: list[tuple[str, str]] = []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            text = "".join(before_units[i1:i2])
            if text:
                before_segments.append(("eq", text))
                after_segments.append(("eq", text))
        elif tag == "delete":
            text = "".join(before_units[i1:i2])
            if text:
                before_segments.append(("del", text))
        elif tag == "insert":
            text = "".join(after_units[j1:j2])
            if text:
                after_segments.append(("add", text))
        elif tag == "replace":
            deleted = "".join(before_units[i1:i2])
            added = "".join(after_units[j1:j2])
            if deleted:
                before_segments.append(("del", deleted))
            if added:
                after_segments.append(("add", added))
    return before_segments, after_segments


def _json_bytes(payload: dict[str, Any]) -> bytes:
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def _write_zip_atomically(package_path: Path, replacements: dict[str, bytes]) -> None:
    package_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = package_path.with_suffix(package_path.suffix + ".tmp")
    if package_path.exists():
        with zipfile.ZipFile(package_path, "r") as src, zipfile.ZipFile(
            temp_path, "w", compression=zipfile.ZIP_DEFLATED
        ) as dst:
            existing = set()
            for info in src.infolist():
                name = info.filename
                existing.add(name)
                replacement = replacements.get(name)
                if replacement is None:
                    dst.writestr(info, src.read(name))
                else:
                    dst.writestr(name, replacement)
            for name, payload in replacements.items():
                if name not in existing:
                    dst.writestr(name, payload)
    else:
        with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as dst:
            for name, payload in replacements.items():
                dst.writestr(name, payload)
    temp_path.replace(package_path)


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha256_file(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as stream:
        while True:
            chunk = stream.read(1024 * 1024)
            if not chunk:
                break
            hasher.update(chunk)
    return hasher.hexdigest()


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()
